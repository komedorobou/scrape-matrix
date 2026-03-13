import streamlit as st
import requests
from bs4 import BeautifulSoup
import time
import random
import re
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import os as _os
import gc as _gc
import sys as _sys
try:
    import cloudscraper
    _scraper = cloudscraper.create_scraper(browser={'browser': 'chrome', 'platform': 'windows', 'mobile': False})
except Exception:
    _scraper = None
# curl_cffi（ブラウザTLS指紋を再現してCloudflare突破）
_curl_session = None
def _get_curl_session():
    """curl_cffi のセッションを取得"""
    global _curl_session
    if _curl_session is not None:
        return _curl_session
    try:
        from curl_cffi.requests import Session
        _curl_session = Session(impersonate="chrome131")
        return _curl_session
    except Exception:
        return None
def _fetch_with_curl(url, timeout=30):
    """curl_cffiでページを取得。成功時は(html, status_code)、失敗時はNone"""
    s = _get_curl_session()
    if s is None:
        return None
    try:
        res = s.get(url, timeout=timeout)
        if res.status_code == 200:
            return res.text
        return None
    except Exception:
        return None
# undetected-chromedriver（ローカルChrome用）
_uc_driver = None
def _get_uc_driver():
    """undetected-chromedriver のドライバーを取得（シングルトン）"""
    global _uc_driver
    if _uc_driver is not None:
        return _uc_driver
    try:
        import undetected_chromedriver as uc
        options = uc.ChromeOptions()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        _uc_driver = uc.Chrome(options=options, version_main=None)
        _uc_driver.set_page_load_timeout(30)
        return _uc_driver
    except Exception:
        return None
def _restart_uc_driver():
    """Chromeドライバーを再起動（メモリリーク対策）"""
    global _uc_driver
    if _uc_driver is not None:
        try:
            _uc_driver.quit()
        except Exception:
            pass
        _uc_driver = None
        _gc.collect()
def _fetch_with_chrome(url, timeout=30):
    """Chromeでページを取得。成功時はHTMLテキスト、失敗時はNone"""
    driver = _get_uc_driver()
    if driver is None:
        return None
    try:
        driver.get(url)
        time.sleep(2)  # JSレンダリング待ち
        return driver.page_source
    except Exception:
        return None
# ページ設定
st.set_page_config(
    page_title="ブランドECスクレイパー",
    page_icon="👜",
    layout="wide",
    initial_sidebar_state="expanded"
)
# viewportメタタグ（スマホでのズーム・スケール制御）
st.markdown('<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">', unsafe_allow_html=True)
# カスタムCSS（グラスモーフィズム＋アニメーション）
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&display=swap');
    /* Material Iconsの文字化け対策 */
    .material-symbols-outlined,
    .material-symbols-rounded,
    [class*="material-symbols"],
    span[data-testid="stIconMaterial"] {
        font-size: 0 !important;
        visibility: hidden !important;
        display: none !important;
    }
    /* サイドバーのラジオボタンを2列グリッドに */
    [data-testid="stSidebar"] [role="radiogroup"] {
        display: grid !important;
        grid-template-columns: 1fr 1fr !important;
        gap: 0.25rem 0.5rem !important;
        padding: 12px 10px !important;
    }
    [data-testid="stSidebar"] [role="radiogroup"] label {
        white-space: nowrap !important;
        font-size: 0.9rem !important;
    }
    /* サイドバー開閉ボタン */
    /* ヘッダーバー（上部の暗い部分）を淡く */
    header[data-testid="stHeader"] {
        background: linear-gradient(135deg, rgba(200, 220, 230, 0.9), rgba(220, 200, 220, 0.9)) !important;
        backdrop-filter: blur(10px) !important;
    }
    header[data-testid="stHeader"] * {
        color: #3a5a5a !important;
    }
    /* ツールバー（右上のメニュー） */
    [data-testid="stToolbar"] {
        background: transparent !important;
    }
    [data-testid="stToolbar"] button {
        color: #4a6a6a !important;
    }
    /* デコレーション（上部のライン）も淡く */
    [data-testid="stDecoration"] {
        background: linear-gradient(90deg, #8ac8d8, #c8a8c8, #a8c8b8) !important;
    }
    /* アニメーション定義 */
    @keyframes gradientShift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    @keyframes pulse {
        0%, 100% { transform: scale(1); }
        50% { transform: scale(1.02); }
    }
    @keyframes shimmer {
        0% { background-position: -200% 0; }
        100% { background-position: 200% 0; }
    }
    /* メイン背景（アニメーショングラデーション） */
    .stApp {
        background: linear-gradient(-45deg, #7ab8a8, #8ac8d8, #a8b8d8, #c8a8c8, #d8a8b8, #b8c8a8);
        background-size: 400% 400%;
        animation: gradientShift 15s ease infinite;
    }
    /* サイドバー（グラスモーフィズム）- 幅を広げる */
    [data-testid="stSidebar"] {
        background: rgba(255, 255, 255, 0.25) !important;
        backdrop-filter: blur(20px) !important;
        border-right: 1px solid rgba(255, 255, 255, 0.3) !important;
        min-width: 350px !important;
        width: 350px !important;
    }
    [data-testid="stSidebar"] > div:first-child {
        width: 350px !important;
        padding: 2rem 1.5rem !important;
    }
    [data-testid="stSidebar"] * {
        color: #2a3a3a !important;
    }
    /* サイドバーのヘッダー */
    [data-testid="stSidebar"] h3 {
        font-size: 1.3rem !important;
        font-weight: 700 !important;
        margin-bottom: 1rem !important;
        padding-bottom: 0.5rem !important;
        border-bottom: 2px solid rgba(138, 200, 216, 0.5) !important;
    }
    /* メインコンテンツエリア（グラスカード） */
    .main .block-container {
        background: rgba(255, 255, 255, 0.4);
        border-radius: 24px;
        padding: 2rem;
        backdrop-filter: blur(20px);
        box-shadow:
            0 8px 32px rgba(0, 0, 0, 0.1),
            inset 0 0 0 1px rgba(255, 255, 255, 0.5);
        border: 1px solid rgba(255, 255, 255, 0.3);
        animation: fadeInUp 0.6s ease-out;
    }
    /* タイトル */
    h1 {
        color: #2a4a4a !important;
        font-family: 'Noto Sans JP', sans-serif !important;
        font-weight: 700 !important;
        letter-spacing: 2px;
        background: linear-gradient(135deg, #4a8a8a, #6a9ab8, #8a7aa8);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    /* サブテキスト */
    .stMarkdown, p, span, label {
        color: #3a5a5a !important;
        font-family: 'Noto Sans JP', sans-serif !important;
    }
    /* 入力フィールド（グラス風） */
    .stTextInput > div > div > input {
        background: rgba(255, 255, 255, 0.85) !important;
        border: 2px solid rgba(138, 200, 216, 0.4) !important;
        border-radius: 12px !important;
        color: #1a3a3a !important;
        font-size: 18px !important;
        font-weight: 500 !important;
        padding: 14px 18px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08) !important;
    }
    .stTextInput > div > div > input::placeholder {
        color: #7a9a9a !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: #6ab8d8 !important;
        background: rgba(255, 255, 255, 0.95) !important;
        box-shadow: 0 0 20px rgba(138, 200, 216, 0.5), 0 4px 12px rgba(0, 0, 0, 0.1) !important;
        transform: translateY(-2px);
    }
    /* 入力フィールドのラベル */
    .stTextInput label {
        font-size: 15px !important;
        font-weight: 600 !important;
        color: #2a4a4a !important;
        margin-bottom: 6px !important;
    }
    /* セレクトボックス */
    .stSelectbox > div > div {
        background: rgba(255, 255, 255, 0.85) !important;
        border: 2px solid rgba(138, 200, 216, 0.4) !important;
        border-radius: 12px !important;
        transition: all 0.3s ease !important;
    }
    .stSelectbox > div > div:hover {
        box-shadow: 0 4px 16px rgba(0, 0, 0, 0.1) !important;
        border-color: #6ab8d8 !important;
    }
    .stSelectbox label {
        font-size: 15px !important;
        font-weight: 600 !important;
        color: #2a4a4a !important;
    }
    /* スライダー */
    .stSlider > div > div > div > div {
        background: linear-gradient(90deg, #8ac8d8, #a8d8e8) !important;
        border-radius: 10px !important;
    }
    /* メインボタン（光沢アニメーション） */
    .stButton > button {
        background: linear-gradient(135deg, #8ac8d8 0%, #7ab8c8 50%, #8ac8d8 100%) !important;
        background-size: 200% 200% !important;
        color: #1a3a3a !important;
        border: none !important;
        border-radius: 16px !important;
        padding: 14px 28px !important;
        font-weight: 700 !important;
        font-size: 16px !important;
        font-family: 'Noto Sans JP', sans-serif !important;
        box-shadow:
            0 4px 15px rgba(138, 200, 216, 0.4),
            inset 0 1px 0 rgba(255, 255, 255, 0.4) !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        position: relative;
        overflow: hidden;
    }
    .stButton > button::before {
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.4), transparent);
        transition: left 0.5s ease;
    }
    .stButton > button:hover {
        transform: translateY(-3px) scale(1.02) !important;
        box-shadow:
            0 8px 25px rgba(138, 200, 216, 0.5),
            inset 0 1px 0 rgba(255, 255, 255, 0.5) !important;
        background-position: 100% 100% !important;
    }
    .stButton > button:hover::before {
        left: 100%;
    }
    .stButton > button:active {
        transform: translateY(-1px) scale(0.98) !important;
    }
    /* ダウンロードボタン */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #7ac8a8 0%, #6ab898 100%) !important;
        color: #1a3a2a !important;
        border: none !important;
        border-radius: 16px !important;
        font-weight: 700 !important;
        font-family: 'Noto Sans JP', sans-serif !important;
        box-shadow: 0 4px 15px rgba(106, 184, 152, 0.4) !important;
        transition: all 0.3s ease !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-3px) !important;
        box-shadow: 0 8px 25px rgba(106, 184, 152, 0.5) !important;
    }
    /* プログレスバー（光沢） */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #7ab8c8, #8ac8d8, #9ad8e8, #8ac8d8) !important;
        background-size: 200% 100% !important;
        animation: shimmer 2s linear infinite !important;
        border-radius: 10px !important;
    }
    /* メトリクスカード */
    [data-testid="stMetricValue"] {
        color: #3a8a7a !important;
        font-size: 2.2rem !important;
        font-weight: 700 !important;
        font-family: 'Noto Sans JP', sans-serif !important;
    }
    [data-testid="stMetricLabel"] {
        color: #5a7a7a !important;
        font-weight: 500 !important;
    }
    /* データフレーム（グラスカード） */
    .stDataFrame {
        background: rgba(255, 255, 255, 0.5) !important;
        border-radius: 16px !important;
        overflow: hidden !important;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08) !important;
        border: 1px solid rgba(255, 255, 255, 0.6) !important;
    }
    /* アラートメッセージ（グラス風） */
    .stSuccess {
        background: rgba(122, 200, 168, 0.25) !important;
        backdrop-filter: blur(10px) !important;
        border-left: 4px solid #6ac898 !important;
        border-radius: 12px !important;
        animation: fadeInUp 0.4s ease-out !important;
    }
    .stInfo {
        background: rgba(138, 200, 216, 0.25) !important;
        backdrop-filter: blur(10px) !important;
        border-left: 4px solid #7ac8e8 !important;
        border-radius: 12px !important;
        animation: fadeInUp 0.4s ease-out !important;
    }
    .stWarning {
        background: rgba(232, 200, 140, 0.25) !important;
        backdrop-filter: blur(10px) !important;
        border-left: 4px solid #e8c878 !important;
        border-radius: 12px !important;
    }
    .stError {
        background: rgba(232, 160, 160, 0.25) !important;
        backdrop-filter: blur(10px) !important;
        border-left: 4px solid #e89898 !important;
        border-radius: 12px !important;
    }
    /* エクスパンダー */
    .streamlit-expanderHeader {
        background: rgba(255, 255, 255, 0.4) !important;
        border-radius: 12px !important;
        color: #3a5a5a !important;
        transition: all 0.3s ease !important;
    }
    .streamlit-expanderHeader:hover {
        background: rgba(255, 255, 255, 0.6) !important;
    }
    /* ヘッダー */
    h2, h3 {
        color: #3a6a6a !important;
        font-family: 'Noto Sans JP', sans-serif !important;
    }
    /* ラジオボタン（カード風） */
    .stRadio > div {
        background: rgba(255, 255, 255, 0.7) !important;
        border-radius: 16px !important;
        padding: 16px !important;
        border: 2px solid rgba(138, 200, 216, 0.3) !important;
        transition: all 0.3s ease !important;
    }
    .stRadio > div:hover {
        background: rgba(255, 255, 255, 0.85) !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1) !important;
        border-color: #8ac8d8 !important;
    }
    .stRadio label {
        font-size: 15px !important;
        font-weight: 600 !important;
        color: #2a4a4a !important;
    }
    .stRadio > div > div > label {
        font-size: 16px !important;
        font-weight: 500 !important;
        padding: 8px 12px !important;
        border-radius: 10px !important;
        transition: all 0.2s ease !important;
    }
    .stRadio > div > div > label:hover {
        background: rgba(138, 200, 216, 0.2) !important;
    }
    /* キャプション（ヘルプテキスト） */
    .stCaption, small, .element-container small {
        font-size: 13px !important;
        color: #4a6a6a !important;
        background: rgba(255, 255, 255, 0.5) !important;
        padding: 8px 12px !important;
        border-radius: 8px !important;
        display: inline-block !important;
        margin-top: 4px !important;
    }
    /* ディバイダー */
    hr {
        border: none !important;
        height: 2px !important;
        background: linear-gradient(90deg, transparent, rgba(138, 200, 216, 0.5), transparent) !important;
        margin: 1.5rem 0 !important;
    }
    /* スライダーのラベル */
    .stSlider label {
        font-size: 15px !important;
        font-weight: 600 !important;
        color: #2a4a4a !important;
    }
    /* スライダーの値表示 */
    .stSlider [data-testid="stTickBarMin"],
    .stSlider [data-testid="stTickBarMax"] {
        font-weight: 600 !important;
        color: #3a5a5a !important;
    }
    /* スクロールバー */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    ::-webkit-scrollbar-track {
        background: rgba(255, 255, 255, 0.2);
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb {
        background: rgba(138, 200, 216, 0.5);
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: rgba(138, 200, 216, 0.7);
    }
    /* ヘルプアイコン（？マーク）を非表示 */
    .stTooltipIcon {
        display: none !important;
    }
    /* ===== モバイル対応（768px以下） ===== */
    @media (max-width: 768px) {
        /* viewportメタタグ的な振る舞い */
        html {
            -webkit-text-size-adjust: 100%;
        }
        /* サイドバー: 全画面幅で開く */
        [data-testid="stSidebar"] {
            min-width: 100vw !important;
            width: 100vw !important;
            z-index: 999 !important;
        }
        [data-testid="stSidebar"] > div:first-child {
            width: 100vw !important;
            padding: 1rem 1.2rem !important;
        }
        /* サイドバー開閉ボタンを大きく・見やすく */
        button[kind="header"] {
            display: block !important;
            min-width: 44px !important;
            min-height: 44px !important;
            font-size: 1.5rem !important;
            z-index: 1000 !important;
        }
        /* サイドバーのラジオボタンを1列に */
        [data-testid="stSidebar"] [role="radiogroup"] {
            grid-template-columns: 1fr !important;
        }
        /* メインコンテンツ: padding/角丸を縮小 */
        .main .block-container {
            padding: 1rem 0.8rem !important;
            border-radius: 12px !important;
            margin: 0.3rem !important;
            max-width: 100% !important;
        }
        /* 横スクロール防止 */
        .main, .block-container, .stApp {
            max-width: 100vw !important;
            overflow-x: hidden !important;
        }
        /* タイトルを小さく */
        h1 {
            font-size: 1.3rem !important;
            letter-spacing: 0px !important;
        }
        h2, h3 {
            font-size: 1.05rem !important;
        }
        /* メトリクスカードを小さく */
        [data-testid="stMetricValue"] {
            font-size: 1.3rem !important;
        }
        [data-testid="stMetricLabel"] {
            font-size: 0.75rem !important;
        }
        /* 入力フィールド: 16px未満だとiOSでズームされる */
        .stTextInput > div > div > input {
            font-size: 16px !important;
            padding: 12px 14px !important;
            border-radius: 10px !important;
        }
        /* ボタンをタップしやすく（44px以上推奨） */
        .stButton > button {
            padding: 14px 20px !important;
            font-size: 16px !important;
            border-radius: 12px !important;
            min-height: 50px !important;
            width: 100% !important;
        }
        .stDownloadButton > button {
            min-height: 50px !important;
            border-radius: 12px !important;
            width: 100% !important;
        }
        /* セレクトボックス・マルチセレクト */
        .stSelectbox > div > div,
        .stMultiSelect > div > div {
            min-height: 44px !important;
            font-size: 16px !important;
        }
        /* multiselect のタグを白背景に（primaryColor CSS変数を上書き） */
        .stMultiSelect [data-baseweb="tag"],
        .stMultiSelect [data-baseweb="tag"][class*="st-"],
        div[data-baseweb="tag"],
        span[data-baseweb="tag"],
        [data-baseweb="tag"][style],
        .stMultiSelect span[data-baseweb="tag"],
        section[data-testid="stSidebar"] [data-baseweb="tag"],
        section[data-testid="stSidebar"] span[data-baseweb="tag"] {
            --primary-color: #FFFFFF;
            font-size: 0.85rem !important;
            padding: 4px 8px !important;
            background: #FFFFFF !important;
            background-color: #FFFFFF !important;
            color: #31333F !important;
            border-radius: 6px !important;
            border: 1px solid #D0D0D0 !important;
        }
        /* BaseWeb Tagのinline style上書き（詳細度最大化） */
        .stApp .stMultiSelect [data-baseweb="tag"],
        .stApp section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"],
        .stApp [data-baseweb="tag"] {
            background: #FFFFFF !important;
            background-color: #FFFFFF !important;
        }
        .stMultiSelect [data-baseweb="tag"] span,
        div[data-baseweb="tag"] span,
        span[data-baseweb="tag"] span,
        [data-baseweb="tag"] span {
            color: #31333F !important;
        }
        .stMultiSelect [data-baseweb="tag"] svg,
        div[data-baseweb="tag"] svg,
        [data-baseweb="tag"] svg {
            color: #555 !important;
            fill: #555 !important;
        }
        /* キャプションを小さく */
        .stCaption, small, .element-container small {
            font-size: 12px !important;
            padding: 6px 10px !important;
        }
        /* データフレームを横スクロールしやすく */
        .stDataFrame {
            border-radius: 8px !important;
        }
        .stDataFrame [data-testid="stDataFrameResizable"] {
            max-height: 350px !important;
        }
        /* アラートメッセージのフォント */
        .stSuccess, .stInfo, .stWarning, .stError {
            font-size: 0.85rem !important;
            border-radius: 8px !important;
            padding: 10px 12px !important;
        }
        /* ラジオボタン */
        .stRadio > div {
            padding: 10px !important;
            border-radius: 10px !important;
        }
        .stRadio > div > div > label {
            font-size: 15px !important;
            padding: 8px 10px !important;
            min-height: 44px !important;
            display: flex !important;
            align-items: center !important;
        }
        /* スライダー: つまみを大きく */
        .stSlider [role="slider"] {
            width: 24px !important;
            height: 24px !important;
        }
        /* エクスパンダー */
        .streamlit-expanderHeader {
            font-size: 0.9rem !important;
            border-radius: 8px !important;
            min-height: 44px !important;
        }
        /* ディバイダー余白縮小 */
        hr {
            margin: 0.8rem 0 !important;
        }
        /* カラム: スマホでは縦並び */
        [data-testid="column"] {
            width: 100% !important;
            flex: 100% !important;
            min-width: 100% !important;
        }
        /* チェックボックスをタップしやすく */
        .stCheckbox label {
            min-height: 44px !important;
            display: flex !important;
            align-items: center !important;
            font-size: 15px !important;
        }
    }
    /* ===== 超小型スマホ（400px以下） ===== */
    @media (max-width: 400px) {
        h1 {
            font-size: 1.1rem !important;
        }
        .main .block-container {
            padding: 0.5rem 0.5rem !important;
            border-radius: 8px !important;
        }
        [data-testid="stMetricValue"] {
            font-size: 1.1rem !important;
        }
        .stButton > button {
            font-size: 15px !important;
            padding: 12px 14px !important;
        }
    }
</style>
""", unsafe_allow_html=True)
# 定数
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
}
# セカスト用セッション（Cookie保持）
_secondstreet_session = None
def _get_secondstreet_session():
    """セカスト用のHTTPセッションを取得（Cookie付き）"""
    global _secondstreet_session
    if _secondstreet_session is not None:
        return _secondstreet_session
    s = requests.Session()
    s.headers.update(HEADERS)
    # まずトップページにアクセスしてCookieを取得
    try:
        s.get("https://www.2ndstreet.jp/", timeout=15)
    except Exception:
        pass
    _secondstreet_session = s
    return s
# EC設定
EC_SITES = {
    "コメ兵": {
        "base_url": "https://komehyo.jp",
        "icon": "🏪"
    },
    "RAGTAG": {
        "base_url": "https://www.ragtag.jp",
        "icon": "👔"
    },
    "トレファク": {
        "base_url": "https://www.trefac.jp",
        "icon": "🏷️"
    },
    "セカスト": {
        "base_url": "https://www.2ndstreet.jp",
        "icon": "🔴"
    }
}
# カテゴリ定義
KOMEHYO_CATEGORIES = {
    "全カテゴリ": "",
    "ブランドバッグ": "brandbag",
    "ブランド財布・小物": "brandwallet-accessories",
    "レディースファッション": "fashion-ladies",
    "メンズファッション": "fashion-mens",
    "時計レディース": "watch-ladies",
    "時計メンズ": "watch-mens",
}
RAGTAG_CATEGORIES = {
    "全カテゴリ": "",
    "メンズ": "men",
    "レディース": "women",
}
TREFAC_CATEGORIES = {
    "全カテゴリ": "",
    "メンズ": "t1",
    "レディース": "t2",
}
SECONDSTREET_CATEGORIES = {
    "全カテゴリ": "",
    "メンズ": "men",
    "レディース": "women",
}
# ===== バックアップ（セッション消失対策） =====
_BACKUP_PATH = _os.path.join("/tmp", "scrape_backup.pkl")

def _save_results_backup(df, brand_name):
    """スクレイピング結果をファイルにバックアップ（セッション消失対策）"""
    try:
        pd.to_pickle({"df": df, "brand": brand_name, "ts": datetime.now()}, _BACKUP_PATH)
    except Exception:
        pass

def _load_results_backup():
    """バックアップから結果を復元（1時間以内のもののみ）"""
    try:
        if _os.path.exists(_BACKUP_PATH):
            data = pd.read_pickle(_BACKUP_PATH)
            if (datetime.now() - data["ts"]).total_seconds() < 3600:
                return data["df"], data["brand"]
    except Exception:
        pass
    return None, None

def _clear_results_backup():
    """バックアップファイルを削除"""
    try:
        if _os.path.exists(_BACKUP_PATH):
            _os.remove(_BACKUP_PATH)
    except Exception:
        pass

# session_stateの初期化
if 'results_df' not in st.session_state:
    st.session_state.results_df = None
if 'brand_name' not in st.session_state:
    st.session_state.brand_name = ""
if 'scraping_done' not in st.session_state:
    st.session_state.scraping_done = False
if 'selected_ec' not in st.session_state:
    st.session_state.selected_ec = "コメ兵"
if 'hinban_mapping_df' not in st.session_state:
    st.session_state.hinban_mapping_df = None
if 'hinban_lookup_done' not in st.session_state:
    st.session_state.hinban_lookup_done = False
if 'hinban_merged_df' not in st.session_state:
    st.session_state.hinban_merged_df = None
if 'wear_catalog_cache' not in st.session_state:
    st.session_state.wear_catalog_cache = {}  # {brand: {brand_no: name}}
if 'brand_suggest_results' not in st.session_state:
    st.session_state.brand_suggest_results = []  # [{name, id, count}, ...]
if 'brand_suggest_keyword' not in st.session_state:
    st.session_state.brand_suggest_keyword = ""
# バックアップからの自動復元（セッション消失時）
if not st.session_state.scraping_done and st.session_state.results_df is None:
    _backup_df, _backup_brand = _load_results_backup()
    if _backup_df is not None:
        st.session_state.results_df = _backup_df
        st.session_state.brand_name = _backup_brand
        st.session_state.scraping_done = True
        st.toast("📂 前回のスクレイピング結果を自動復元しました")
# ===== コメ兵用関数 =====
def komehyo_build_url(brand, category):
    """コメ兵: 検索URLを構築"""
    base_url = EC_SITES["コメ兵"]["base_url"]
    brand_clean = brand.lower().strip().replace(" ", "")
    if category:
        return f"{base_url}/{category}/{brand_clean}/"
    else:
        return f"{base_url}/{brand_clean}/"
def komehyo_get_product_urls(base_url, max_pages, progress_callback=None):
    """コメ兵: 商品URLを全ページから取得（リトライ付き）"""
    urls = []
    page = 1
    consecutive_errors = 0
    while page <= max_pages:
        url = f"{base_url}?page={page}"
        if progress_callback:
            progress_callback(f"📄 ページ {page}/{max_pages} の商品リストを取得中...（{len(urls)}件発見済み）")
        try:
            res = requests.get(url, headers=HEADERS, timeout=30)
            if res.status_code != 200:
                consecutive_errors += 1
                if consecutive_errors >= 3:
                    if progress_callback:
                        progress_callback(f"⚠ 3回連続エラーで中断 / {len(urls)}件取得済み")
                    break
                page += 1
                time.sleep(1)
                continue
            soup = BeautifulSoup(res.text, 'html.parser')
            links = soup.select('a[href*="/product/"]')
            new_urls = []
            for a in links:
                href = a.get('href', '')
                if '/product/' in href:
                    full_url = EC_SITES["コメ兵"]["base_url"] + href if href.startswith('/') else href
                    new_urls.append(full_url)
            new_urls = list(set(new_urls))
            if not new_urls:
                break
            before_count = len(urls)
            urls.extend(new_urls)
            urls = list(set(urls))
            if len(urls) == before_count:
                break
            consecutive_errors = 0
            page += 1
            time.sleep(random.uniform(0.5, 1.0))
        except Exception as e:
            consecutive_errors += 1
            if consecutive_errors >= 3:
                if progress_callback:
                    progress_callback(f"⚠ 3回連続例外で中断 / {len(urls)}件取得済み")
                break
            page += 1
            time.sleep(2)
    return urls
def komehyo_get_product_detail(url):
    """コメ兵: 商品詳細を取得"""
    try:
        res = requests.get(url, headers=HEADERS, timeout=30)
        soup = BeautifulSoup(res.text, 'html.parser')
        data = {"URL": url}
        h1 = soup.find('h1')
        data["商品名"] = h1.get_text(strip=True) if h1 else ""
        price_el = soup.select_one('[class*="price"]')
        if price_el:
            price_text = price_el.get_text()
            match = re.search(r'[￥¥]([\d,]+)', price_text)
            if match:
                data["価格"] = int(match.group(1).replace(',', ''))
        rows = soup.select('tr')
        for row in rows:
            th = row.find('th')
            td = row.find('td')
            if th and td:
                key = th.get_text(strip=True)
                val = td.get_text(strip=True).split('\n')[0].strip()
                if key == "品番型式":
                    data["品番"] = val
                elif key == "アイテム名":
                    data["商品名"] = val  # h1より正確なアイテム名で上書き
                elif key == "通称":
                    data["通称"] = val  # 正式名称（ピーカブー等）
                elif key == "商品ランク":
                    rank_match = re.search(r'(新品|未使用品|中古品[SABC])', val)
                    if rank_match:
                        data["ランク"] = rank_match.group(1)
                elif key == "カラー":
                    data["カラー"] = val
                elif key == "素材":
                    data["素材"] = val
                elif key == "ブランド":
                    data["ブランド"] = val
                elif key == "性別タイプ":
                    data["性別"] = val
                elif key == "参考上代":
                    ref_match = re.search(r'[￥¥]([\d,]+)', val)
                    if ref_match:
                        data["参考上代"] = int(ref_match.group(1).replace(',', ''))
        return data
    except Exception as e:
        return None
# ===== RAGTAG用関数 =====
def ragtag_build_url(brand, category):
    """RAGTAG: 検索URLを構築"""
    base_url = EC_SITES["RAGTAG"]["base_url"]
    # 検索URL形式: https://www.ragtag.jp/search?fr=FENDI
    brand_clean = brand.upper().strip().replace(" ", "")
    return f"{base_url}/search?fr={brand_clean}"
def ragtag_get_product_urls(base_url, max_pages, progress_callback=None):
    """RAGTAG: 商品URLを全ページから取得（リトライ付き）"""
    urls = []
    page = 1
    consecutive_errors = 0
    while page <= max_pages:
        # ページネーション: &page=2, &page=3...
        url = f"{base_url}&page={page}" if page > 1 else base_url
        if progress_callback:
            progress_callback(f"📄 ページ {page}/{max_pages} の商品リストを取得中...（{len(urls)}件発見済み）")
        try:
            res = requests.get(url, headers=HEADERS, timeout=30)
            if res.status_code != 200:
                consecutive_errors += 1
                if consecutive_errors >= 3:
                    if progress_callback:
                        progress_callback(f"⚠ 3回連続エラーで中断 / {len(urls)}件取得済み")
                    break
                page += 1
                time.sleep(1)
                continue
            soup = BeautifulSoup(res.text, 'html.parser')
            # RAGTAGの商品リンクを取得（/item/ を含むリンク）
            all_links = soup.find_all('a', href=True)
            new_urls = []
            for a in all_links:
                href = a.get('href', '')
                if '/item/' in href:
                    if href.startswith('http'):
                        full_url = href
                    elif href.startswith('/'):
                        full_url = EC_SITES["RAGTAG"]["base_url"] + href
                    else:
                        full_url = EC_SITES["RAGTAG"]["base_url"] + '/' + href
                    new_urls.append(full_url)
            new_urls = list(set(new_urls))
            if not new_urls:
                break
            before_count = len(urls)
            urls.extend(new_urls)
            urls = list(set(urls))
            if len(urls) == before_count:
                break
            consecutive_errors = 0
            page += 1
            time.sleep(random.uniform(0.5, 1.0))
        except Exception as e:
            consecutive_errors += 1
            if consecutive_errors >= 3:
                if progress_callback:
                    progress_callback(f"⚠ 3回連続例外で中断 / {len(urls)}件取得済み")
                break
            page += 1
            time.sleep(2)
    return urls
def ragtag_get_product_detail(url):
    """RAGTAG: 商品詳細を取得"""
    try:
        res = requests.get(url, headers=HEADERS, timeout=30)
        soup = BeautifulSoup(res.text, 'html.parser')
        data = {"URL": url}
        # ブランド名
        brand_el = soup.select_one('.item-detail-info__name-brand')
        if brand_el:
            brand_text = brand_el.get_text(strip=True).split('\n')[0].strip()
            data["ブランド"] = brand_text
        # カテゴリ
        cat_el = soup.select_one('.item-detail-info__category-list')
        if cat_el:
            data["カテゴリ"] = cat_el.get_text(strip=True).replace('\n', '').replace(' ', '')
        # 価格
        price_el = soup.select_one('.item-detail-info__price')
        if price_el:
            price_text = price_el.get_text(strip=True).replace(',', '')
            match = re.search(r'(\d+)', price_text)
            if match:
                data["価格"] = int(match.group(1))
        # カラー
        color_el = soup.select_one('.item-detail-info__sku-color-name')
        if color_el:
            data["カラー"] = color_el.get_text(strip=True)
        # コンディション・サイズ（正規表現でテキストから抽出）
        page_text = soup.get_text()
        cond_match = re.search(r'コンディション\s*[:：]\s*(\w+)', page_text)
        if cond_match:
            data["ランク"] = cond_match.group(1)
        size_match = re.search(r'サイズ\s*[:：]\s*([^\s]+)', page_text)
        if size_match:
            data["サイズ"] = size_match.group(1)
        # 商品名（metaのdescriptionから取得）
        meta_desc = soup.select_one('meta[name="description"]')
        if meta_desc:
            desc_content = meta_desc.get('content', '')
            desc_parts = re.split(r'<br>|<BR>', desc_content)
            for part in desc_parts:
                part = part.strip()
                if part and 2 <= len(part) <= 30:
                    if not re.search(r'(の商品|公式|通販|買取|販売|サイト|RAGTAG|ラグタグ|送料)', part):
                        if part not in ["FENDI", "GUCCI", "PRADA", "CHANEL", "HERMES", "CELINE", "LOEWE"]:
                            data["商品名"] = part
                            break
        # 品番（metaのdescriptionから抽出）
        if meta_desc:
            desc_content = meta_desc.get('content', '')
            candidates = re.findall(r'\b[A-Z0-9]{5,10}\b', desc_content)
            for candidate in candidates:
                has_digit = any(c.isdigit() for c in candidate)
                has_alpha = any(c.isalpha() for c in candidate)
                if has_digit and has_alpha and candidate not in ["FENDI", "GUCCI", "PRADA", "CHANEL", "HERMES", "CELINE", "LOEWE"]:
                    data["品番"] = candidate
                    break
        return data
    except Exception as e:
        return None
# ===== トレファク用関数 =====
def trefac_build_url(brand, category):
    """トレファク: 検索URLを構築"""
    base_url = EC_SITES["トレファク"]["base_url"]
    brand_clean = brand.strip()
    if category:
        return f"{base_url}/store/{category}cpsb/?srchword={brand_clean}"
    else:
        return f"{base_url}/store/search_result.html?srchword={brand_clean}"
def trefac_get_product_urls(base_url, max_pages, progress_callback=None):
    """トレファク: 商品URLを全ページから取得（リトライ付き）"""
    urls = []
    page = 1
    consecutive_errors = 0
    while page <= max_pages:
        separator = "&" if "?" in base_url else "?"
        url = f"{base_url}{separator}key={page}"
        if progress_callback:
            progress_callback(f"📄 ページ {page}/{max_pages} の商品リストを取得中...（{len(urls)}件発見済み）")
        try:
            res = requests.get(url, headers=HEADERS, timeout=30)
            if res.status_code != 200:
                consecutive_errors += 1
                if consecutive_errors >= 3:
                    if progress_callback:
                        progress_callback(f"⚠ 3回連続エラー（HTTP {res.status_code}）で中断 / {len(urls)}件取得済み")
                    break
                page += 1
                time.sleep(1)
                continue
            soup = BeautifulSoup(res.text, 'html.parser')
            all_links = soup.find_all('a', href=True)
            new_urls = []
            for a in all_links:
                href = a.get('href', '')
                if re.search(r'/store/\d{10,}/', href):
                    if href.startswith('http'):
                        full_url = href
                    elif href.startswith('/'):
                        full_url = EC_SITES["トレファク"]["base_url"] + href
                    else:
                        full_url = EC_SITES["トレファク"]["base_url"] + '/' + href
                    new_urls.append(full_url)
            new_urls = list(set(new_urls))
            if not new_urls:
                break
            before_count = len(urls)
            urls.extend(new_urls)
            urls = list(set(urls))
            if len(urls) == before_count:
                break
            consecutive_errors = 0
            page += 1
            time.sleep(random.uniform(0.5, 1.0))
        except Exception as e:
            consecutive_errors += 1
            if consecutive_errors >= 3:
                if progress_callback:
                    progress_callback(f"⚠ 3回連続例外で中断 / {len(urls)}件取得済み")
                break
            if progress_callback:
                progress_callback(f"⚠ ページ {page} エラー: {type(e).__name__}（リトライ）")
            page += 1
            time.sleep(2)
    return urls
_trefac_session = None
def _get_trefac_session():
    global _trefac_session
    if _trefac_session is None:
        _trefac_session = requests.Session()
        _trefac_session.headers.update(HEADERS)
    return _trefac_session

# スレッドローカルセッション（並列処理用：スレッドごとにコネクション再利用）
_trefac_thread_local = threading.local()
def _get_trefac_thread_session():
    """スレッドごとのrequests.Sessionを取得（コネクション再利用で高速化）"""
    if not hasattr(_trefac_thread_local, 'session'):
        s = requests.Session()
        s.headers.update(HEADERS)
        # コネクションプールサイズを拡大
        adapter = requests.adapters.HTTPAdapter(pool_connections=10, pool_maxsize=10)
        s.mount('https://', adapter)
        s.mount('http://', adapter)
        _trefac_thread_local.session = s
    return _trefac_thread_local.session

def trefac_get_product_detail(url):
    """トレファク: 商品詳細を取得（並列呼び出し対応・スレッドローカルSession）"""
    try:
        # スレッドローカルSessionでコネクション再利用（5000件で大幅高速化）
        ss = _get_trefac_thread_session()
        res = ss.get(url, timeout=30)
        soup = BeautifulSoup(res.text, 'html.parser')
        data = {"URL": url}
        # 価格
        price_el = soup.select_one('.gdprice_main')
        if price_el:
            price_text = price_el.get_text(strip=True).replace(',', '')
            match = re.search(r'(\d+)', price_text)
            if match:
                data["価格"] = int(match.group(1))
        # 属性テーブルから取得
        attr_rows = soup.select('.gddescription_attr_row')
        for row in attr_rows:
            head = row.select_one('.gddescription_attr_head')
            data_el = row.select_one('.gddescription_attr_data')
            if head and data_el:
                key = head.get_text(strip=True)
                val = data_el.get_text(strip=True).replace('\n', ' ')
                if 'ブランド' in key:
                    data["ブランド"] = val.split()[0] if val else ""
                elif '性別' in key:
                    data["性別"] = val
                elif 'カテゴリ' in key:
                    data["カテゴリ"] = val.replace('>', ' > ').strip()
                elif 'コンディション' in key:
                    stars = val.count('★')
                    data["ランク"] = f"★{stars}"
                elif '付属品' in key:
                    data["付属品"] = val
        # 詳細テーブルから取得
        detail_rows = soup.select('.gddescription_detail_row')
        for row in detail_rows:
            head = row.select_one('.gddescription_detail_head')
            data_el = row.select_one('.gddescription_detail_data')
            if head and data_el:
                key = head.get_text(strip=True)
                val = data_el.get_text(strip=True).replace('\n', ' ')[:100]
                if 'アイテム名' in key:
                    data["商品名"] = val
                elif '型番' in key:
                    data["型番"] = val
                    data["品番"] = val
                elif 'カラー' in key:
                    data["カラー"] = val
                elif '素材' in key:
                    data["素材"] = val
                elif '製造国' in key:
                    data["製造国"] = val
                elif 'サイズ' in key and 'サイズ' not in data:
                    data["サイズ"] = val
        return data
    except Exception as e:
        return None
# ===== セカスト用関数 =====
_secondstreet_cache = {}
def _parse_secondstreet_name(full_name):
    """商品名文字列を 商品名 / 型番 / 素材 / カラー / 商品番号 に分割"""
    if not full_name:
        return full_name, "", "", "", ""
    parts = [p.strip() for p in full_name.split('/') if p.strip()]
    if not parts:
        return full_name, "", "", "", ""
    product_name = parts[0]
    model_parts = []
    material = ""
    color = ""
    product_number = ""
    known_materials = {'コットン', 'シルク', 'ポリエステル', 'ナイロン', 'ウール', 'レーヨン',
                       'リネン', 'カシミヤ', 'レザー', 'デニム', 'スエード', 'アクリル', 'キュプラ'}
    known_colors_en = {'BLK', 'WHT', 'GRY', 'BLU', 'RED', 'GRN', 'NVY', 'BRW', 'PNK',
                       'YLW', 'ORG', 'PRP', 'BGE', 'IVR', 'GLD', 'SLV', 'CRM'}
    for part in parts[1:]:
        upper = part.upper().strip()
        if not product_number and re.fullmatch(r'\d{5,}', part):
            product_number = part
        elif re.search(r'[A-Za-z]', part) and re.search(r'\d', part) and len(part) >= 4:
            model_parts.append(part)
        elif not material and any(m in part for m in known_materials):
            material = part
        elif not color and upper in known_colors_en:
            color = part
    model = " / ".join(model_parts) if model_parts else ""
    return product_name, model, material, color, product_number
def secondstreet_build_url(brand, category):
    """セカスト: 検索URLを構築"""
    base_url = EC_SITES["セカスト"]["base_url"]
    brand_clean = brand.strip()
    return f"{base_url}/search?keyword={brand_clean}"
def _ss_extract_chrome_cookies():
    """Chromeからクッキーを抽出して返す"""
    cookies = {}
    if _uc_driver is not None:
        try:
            for c in _uc_driver.get_cookies():
                cookies[c['name']] = c['value']
        except Exception:
            pass
    return cookies
def secondstreet_get_product_urls(base_url, max_pages, progress_callback=None, save_callback=None, existing_urls=None):
    """セカスト: 検索ページから商品URL+データを一括取得（リトライ付き・メモリ最適化）
    save_callback: 10ページごとに呼ばれるコールバック(cache_dict) → 中間保存用
    existing_urls: 既存取得済みURL集合（再開時にスキップ用）"""
    global _secondstreet_cache
    _secondstreet_cache = {}
    # 再開モード: 既存URLをキャッシュに入れて重複スキップ
    if existing_urls:
        for u in existing_urls:
            _secondstreet_cache[u] = {"_skip": True}
    urls = []
    page = 1
    consecutive_errors = 0
    # Chrome→Cookie抽出→requests切替でメモリ節約
    _chrome_cookies = {}
    _preferred_method = None  # 1ページ目で成功した方式を記憶
    _cookie_refresh_count = 0  # Cookie再取得の回数制限
    _empty_page_count = 0  # カード0件の連続回数
    _consecutive_skip_pages = 0  # 再開モード: 連続全スキップページ数
    while page <= max_pages:
        url = f"{base_url}&page={page}" if page > 1 else base_url
        if progress_callback:
            progress_callback(f"📄 ページ {page}/{max_pages} の商品リストを取得中...（{len(urls)}件発見済み）")
        try:
            html = None
            method = ""
            # 0) ScraperAPI（APIキーがあれば最優先・確実）
            scraper_api_key = st.session_state.get("scraper_api_key", "")
            if not html and scraper_api_key:
                try:
                    api_url = f"https://api.scraperapi.com?api_key={scraper_api_key}&url={url}&render=false"
                    res = requests.get(api_url, timeout=60)
                    if res.status_code == 200:
                        if '.itemCard' in res.text or 'itemCard' in res.text:
                            html = res.text
                            method = "ScraperAPI"
                        else:
                            api_url2 = (f"https://api.scraperapi.com?api_key={scraper_api_key}"
                                        f"&url={url}&render=true&wait=5000")
                            res2 = requests.get(api_url2, timeout=90)
                            if res2.status_code == 200:
                                html = res2.text
                                method = "ScraperAPI(render)"
                except Exception:
                    pass
            # 2ページ目以降: 前回成功した軽量方式を優先（Chrome回避）
            if not html and _preferred_method and page > 1:
                if _preferred_method == "curl_cffi":
                    html = _fetch_with_curl(url)
                    if html:
                        method = "curl_cffi"
                elif _preferred_method == "cloudscraper" and _scraper:
                    try:
                        res = _scraper.get(url, headers=HEADERS, timeout=30)
                        if res.status_code == 200:
                            html = res.text
                            method = "cloudscraper"
                    except Exception:
                        pass
                elif _preferred_method == "requests+cookies" and _chrome_cookies:
                    ss = _get_secondstreet_session()
                    try:
                        ss.cookies.update(_chrome_cookies)
                        res = ss.get(url, timeout=30)
                        if res.status_code == 200 and ('itemCard' in res.text):
                            html = res.text
                            method = "requests+cookies"
                    except Exception:
                        pass
            # 1) Chrome（1ページ目 or Cookie切れ時のリフレッシュ用、最大3回）
            if not html and (page == 1 or _cookie_refresh_count < 3):
                html = _fetch_with_chrome(url)
                if html:
                    method = "Chrome" if page == 1 else "Chrome(refresh)"
                    # Cookieを抽出
                    _chrome_cookies = _ss_extract_chrome_cookies()
                    if page > 1:
                        _cookie_refresh_count += 1
                        if progress_callback:
                            progress_callback(f"🔄 Cookie再取得（{_cookie_refresh_count}/3）/ {len(urls)}件取得済み")
                    # Cookie取得後すぐChrome解放
                    _restart_uc_driver()
                    if _chrome_cookies:
                        _preferred_method = "requests+cookies"
            # 2) curl_cffi（TLS指紋偽装）
            if not html:
                html = _fetch_with_curl(url)
                if html:
                    method = "curl_cffi"
            # 3) cloudscraper
            if not html and _scraper:
                try:
                    res = _scraper.get(url, headers=HEADERS, timeout=30)
                    if res.status_code == 200:
                        html = res.text
                        method = "cloudscraper"
                except Exception:
                    pass
            # 4) requests Session（Cookie保持で403回避）
            if not html:
                ss = _get_secondstreet_session()
                try:
                    if _chrome_cookies:
                        ss.cookies.update(_chrome_cookies)
                    res = ss.get(url, timeout=30)
                    if res.status_code == 200:
                        html = res.text
                        method = "requests"
                    else:
                        if progress_callback:
                            progress_callback(f"⚠ ページ {page}: HTTP {res.status_code}（リトライ）")
                except Exception:
                    pass
            if not html:
                consecutive_errors += 1
                if consecutive_errors >= 5:
                    if progress_callback:
                        progress_callback(f"⚠ 5回連続HTML取得失敗で中断 / {len(urls)}件取得済み")
                    break
                if progress_callback:
                    progress_callback(f"⚠ ページ {page}: HTML取得失敗（{consecutive_errors}/5）リトライ中...")
                # 段階的バックオフ（ページは進めずリトライ）
                time.sleep(min(2 * consecutive_errors, 10))
                continue
            # 1ページ目で成功した方式を記憶（Chrome以外）
            if page == 1:
                if method in ("curl_cffi", "cloudscraper", "ScraperAPI", "ScraperAPI(render)"):
                    _preferred_method = method
                elif method in ("Chrome", "Chrome(refresh)") and _chrome_cookies:
                    _preferred_method = "requests+cookies"
                if progress_callback:
                    progress_callback(f"🔧 取得方式: {method}" + (f" → 2p以降: {_preferred_method}" if _preferred_method else ""))
            soup = BeautifulSoup(html, 'html.parser')
            cards = soup.select('.itemCard')
            if not cards:
                _empty_page_count += 1
                del soup, html
                if _empty_page_count >= 3:
                    # 3回連続カード0件で本当にデータ終了
                    if progress_callback:
                        progress_callback(f"⚠ 3回連続カード0件でデータ終了 / {len(urls)}件取得済み")
                    break
                # 1-2回目はスキップして次ページへ（一時的なブロックの可能性）
                if progress_callback:
                    progress_callback(f"⚠ ページ {page}: カード0件（{_empty_page_count}/3）スキップ...")
                page += 1
                time.sleep(2)
                continue
            _empty_page_count = 0  # カード取得成功でリセット
            new_count = 0
            for card in cards:
                link = card.select_one('a.itemCard_inner')
                if not link:
                    continue
                href = link.get('href', '')
                if '/goods/detail/goodsId/' not in href:
                    continue
                if href.startswith('http'):
                    product_url = href
                elif href.startswith('/'):
                    product_url = EC_SITES["セカスト"]["base_url"] + href
                else:
                    product_url = EC_SITES["セカスト"]["base_url"] + '/' + href
                if product_url in _secondstreet_cache:
                    continue
                data = {"URL": product_url}
                brand_el = card.select_one('.itemCard_brand')
                if brand_el:
                    data["ブランド"] = brand_el.get_text(strip=True)
                name_el = card.select_one('.itemCard_name')
                full_name = name_el.get_text(strip=True) if name_el else ""
                product_name, model, material, color, product_number = _parse_secondstreet_name(full_name)
                data["商品名"] = product_name
                if model:
                    data["型番"] = model
                if material:
                    data["素材"] = material
                if color:
                    data["カラー"] = color
                if product_number:
                    data["品番"] = product_number
                elif model:
                    data["品番"] = model.split(" / ")[0]
                size_el = card.select_one('.itemCard_size')
                if size_el:
                    size_text = size_el.get_text(strip=True).replace('サイズ', '').strip()
                    if size_text and size_text != '--':
                        data["サイズ"] = size_text
                status_el = card.select_one('.itemCard_status')
                if status_el:
                    status_text = status_el.get_text(strip=True)
                    rank_match = re.search(r'中古([A-Za-z])', status_text)
                    if rank_match:
                        data["ランク"] = f"中古{rank_match.group(1).upper()}"
                    elif '未使用' in status_text:
                        data["ランク"] = "未使用品"
                price_el = card.select_one('.itemCard_price')
                if price_el:
                    price_text = price_el.get_text(strip=True).replace('¥', '').replace('￥', '').replace(',', '')
                    price_match = re.search(r'(\d+)', price_text)
                    if price_match:
                        data["価格"] = int(price_match.group(1))
                _secondstreet_cache[product_url] = data
                urls.append(product_url)
                new_count += 1
            if new_count == 0:
                del soup, html, cards
                if existing_urls:
                    # 再開モード: 既存URLのページはスキップして続行
                    _consecutive_skip_pages += 1
                    if _consecutive_skip_pages >= 20:
                        # 20ページ連続で全部既存→もう新規はないと判断
                        if progress_callback:
                            progress_callback(f"ℹ️ 20ページ連続スキップ → 新規データなしで終了 / {len(urls)}件追加取得済み")
                        break
                    page += 1
                    time.sleep(random.uniform(0.3, 0.5))
                    continue
                break
            # 新規データあり→連続スキップカウンタリセット
            if existing_urls:
                _consecutive_skip_pages = 0
            # メモリ解放（毎ページ）
            del soup, html, cards
            if page % 10 == 0:
                _gc.collect()
                # 10ページごとに中間保存（OOM kill対策）
                if save_callback and _secondstreet_cache:
                    save_callback(_secondstreet_cache)
                if progress_callback:
                    progress_callback(f"📄 {len(urls)}件取得済み（保存+メモリ解放済み）")
            consecutive_errors = 0
            page += 1
            time.sleep(random.uniform(0.5, 1.0))
        except MemoryError:
            # メモリ不足は即座にブレイク（持ってるデータを救う）
            _gc.collect()
            _restart_uc_driver()
            if progress_callback:
                progress_callback(f"⚠ メモリ不足で中断 / {len(urls)}件取得済み")
            break
        except Exception as e:
            consecutive_errors += 1
            if consecutive_errors >= 5:
                if progress_callback:
                    progress_callback(f"⚠ 5回連続例外で中断 / {len(urls)}件取得済み: {type(e).__name__}")
                break
            if progress_callback:
                progress_callback(f"⚠ ページ {page} 例外: {type(e).__name__}（{consecutive_errors}/5）リトライ中...")
            # 段階的バックオフ
            time.sleep(min(2 * consecutive_errors, 10))
    # 終了時にChrome確実に解放
    _restart_uc_driver()
    return urls
def secondstreet_get_product_detail(url):
    """セカスト: キャッシュから商品データを返す（検索ページで取得済み）"""
    data = _secondstreet_cache.get(url)
    if data and "_skip" in data:
        return None  # 再開モードで既存URLはスキップ
    return data
def get_products_parallel(urls, get_detail_func, max_workers=10, progress_callback=None, batch_size=200):
    """並列処理で商品詳細を取得（バッチ処理対応）"""
    results = []
    total = len(urls)
    completed = 0
    # バッチに分割して処理（メモリ節約＋安定性向上）
    for batch_start in range(0, total, batch_size):
        batch_urls = urls[batch_start:batch_start + batch_size]
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(get_detail_func, url): url for url in batch_urls}
            for future in as_completed(futures):
                try:
                    data = future.result(timeout=60)
                except Exception:
                    data = None
                if data:
                    results.append(data)
                completed += 1
                if progress_callback:
                    progress_callback(completed, total)
    return results
# ===== 件数チェック（1ページ目だけ取得して概算） =====
def _check_count_komehyo(brand, category):
    """コメ兵: 検索結果件数を概算取得"""
    url = komehyo_build_url(brand, category)
    try:
        res = requests.get(url, headers=HEADERS, timeout=15)
        if res.status_code != 200:
            return None, url
        soup = BeautifulSoup(res.text, 'html.parser')
        # 件数テキストを探す（例: "123件"）
        text = soup.get_text()
        m = re.search(r'([\d,]+)\s*件', text)
        if m:
            return int(m.group(1).replace(',', '')), url
        # フォールバック: 商品リンク数をカウント
        links = soup.select('a[href*="/product/"]')
        count = len(set(a.get('href', '') for a in links))
        return (count, url) if count else (0, url)
    except Exception:
        return None, url

def _check_count_ragtag(brand, category):
    """RAGTAG: 検索結果件数を概算取得"""
    url = ragtag_build_url(brand, category)
    try:
        res = requests.get(url, headers=HEADERS, timeout=15)
        if res.status_code != 200:
            return None, url
        soup = BeautifulSoup(res.text, 'html.parser')
        text = soup.get_text()
        m = re.search(r'([\d,]+)\s*件', text)
        if m:
            return int(m.group(1).replace(',', '')), url
        links = [a for a in soup.find_all('a', href=True) if '/item/' in a.get('href', '')]
        count = len(set(a.get('href', '') for a in links))
        return (count, url) if count else (0, url)
    except Exception:
        return None, url

def _check_count_trefac(brand, category):
    """トレファク: 検索結果件数を概算取得"""
    url = trefac_build_url(brand, category)
    try:
        res = requests.get(url, headers=HEADERS, timeout=15)
        if res.status_code != 200:
            return None, url
        soup = BeautifulSoup(res.text, 'html.parser')
        text = soup.get_text()
        m = re.search(r'([\d,]+)\s*件', text)
        if m:
            return int(m.group(1).replace(',', '')), url
        links = [a for a in soup.find_all('a', href=True) if re.search(r'/store/\d{10,}/', a.get('href', ''))]
        count = len(set(a.get('href', '') for a in links))
        return (count, url) if count else (0, url)
    except Exception:
        return None, url

def _parse_secondstreet_count(html, method):
    """セカスト: HTMLから件数を抽出（複数パターン対応）"""
    soup = BeautifulSoup(html, 'html.parser')
    title = soup.select_one('title')
    title_text = title.get_text(strip=True) if title else "(no title)"
    # 1) JSON内の件数を探す（最大値を優先）
    json_counts = {}
    for script in soup.select('script'):
        script_text = script.string or ""
        for key in ['totalCount', 'total', 'itemCount', 'hit_count', 'hitCount',
                     'totalItems', 'total_count', 'numFound', 'nbHits', 'totalHits']:
            for m in re.finditer(rf'["\']?{key}["\']?\s*[:=]\s*(\d+)', script_text):
                val = int(m.group(1))
                if val > json_counts.get(key, 0):
                    json_counts[key] = val
    if json_counts:
        best_key = max(json_counts, key=json_counts.get)
        best_val = json_counts[best_key]
        if best_val > 20:
            return best_val, f"{method} / json:{best_key}={best_val} / title={title_text}"
    # 2) テキスト中の「XX件」— 全て探して最大値
    text = soup.get_text()
    max_count = 0
    max_match = ""
    for pat in [r'([\d,]+)\s*件', r'全\s*([\d,]+)', r'(\d+)\s*(?:items|results|products)']:
        for m in re.finditer(pat, text):
            count = int(m.group(1).replace(',', ''))
            if count > max_count:
                max_count = count
                max_match = m.group(0)
    if max_count > 20:
        return max_count, f"{method} / text:'{max_match}'(max) / title={title_text}"
    # 3) ページネーションから推定（最終ページ × 1ページの商品数）
    cards = soup.select('.itemCard')
    cards_per_page = len(cards) if cards else 0
    last_page = 1
    for a in soup.find_all('a', href=True):
        m = re.search(r'[?&]page=(\d+)', a.get('href', ''))
        if m:
            pg = int(m.group(1))
            if pg > last_page:
                last_page = pg
    for el in soup.select('[class*="pager"], [class*="pagin"], [class*="Pager"], nav'):
        for m in re.finditer(r'\b(\d+)\b', el.get_text()):
            pg = int(m.group(1))
            if 1 < pg < 1000 and pg > last_page:
                last_page = pg
    if last_page > 1 and cards_per_page > 0:
        estimated = last_page * cards_per_page
        return estimated, f"{method} / pagination:{last_page}p×{cards_per_page}cards / title={title_text}"
    # 4) カード数のみ
    if cards_per_page > 0:
        return cards_per_page, f"{method} / cards={cards_per_page}(1p) / title={title_text}"
    for selector in ['[class*="itemCard"]', '[class*="item-card"]', '.item_list li']:
        found = soup.select(selector)
        if found:
            return len(found), f"{method} / {selector}={len(found)} / title={title_text}"
    # 5) テキストの件数（小さくても返す）
    if max_count > 0:
        return max_count, f"{method} / text:'{max_match}'(small) / title={title_text}"
    return 0, f"{method} / HTML={len(html)}文字 / title={title_text} / 何も見つからず"

def _fetch_secondstreet_html(url, scraper_api_key=""):
    """セカスト: URLからHTMLを取得（共通フォールバック）"""
    html = None
    if scraper_api_key:
        try:
            api_url = (f"https://api.scraperapi.com?api_key={scraper_api_key}"
                       f"&url={url}&render=true&wait=5000")
            res = requests.get(api_url, timeout=90)
            if res.status_code == 200:
                html = res.text
        except Exception:
            pass
        if not html:
            try:
                api_url = (f"https://api.scraperapi.com?api_key={scraper_api_key}"
                           f"&url={url}&render=false")
                res = requests.get(api_url, timeout=60)
                if res.status_code == 200:
                    html = res.text
            except Exception:
                pass
    if not html:
        html = _fetch_with_chrome(url)
    if not html:
        html = _fetch_with_curl(url)
    if not html and _scraper:
        try:
            res = _scraper.get(url, headers=HEADERS, timeout=15)
            if res.status_code == 200:
                html = res.text
        except Exception:
            pass
    if not html:
        try:
            ss = _get_secondstreet_session()
            res = ss.get(url, timeout=30)
            if res.status_code == 200:
                html = res.text
        except Exception:
            pass
    return html

def _suggest_brands_secondstreet(keyword, scraper_api_key=""):
    """セカスト: ブランド名サジェスト（検索ページのブランドフィルタから取得）"""
    if not keyword or len(keyword) < 2:
        return []
    url = f"https://www.2ndstreet.jp/search?keyword={keyword.strip()}"
    html = _fetch_secondstreet_html(url, scraper_api_key)
    if not html:
        return []
    soup = BeautifulSoup(html, 'html.parser')
    brands = []
    # 1) __NUXT__ / script JSON内のブランドデータを探す
    for script in soup.select('script'):
        text = script.string or ""
        # brandName/brandId パターン
        for m in re.finditer(
            r'\{[^{}]*"brandName"\s*:\s*"([^"]+)"[^{}]*"brandId"\s*:\s*"?(\d+)"?[^{}]*\}',
            text
        ):
            name, bid = m.group(1), m.group(2)
            # 件数も取れたら取得
            count_m = re.search(r'"count"\s*:\s*(\d+)', m.group(0))
            count = int(count_m.group(1)) if count_m else None
            brands.append({"name": name, "id": bid, "count": count})
        # brand配列パターン（逆順: id, name）
        for m in re.finditer(
            r'\{[^{}]*"id"\s*:\s*"?(\d+)"?[^{}]*"name"\s*:\s*"([^"]+)"[^{}]*\}',
            text
        ):
            bid, name = m.group(1), m.group(2)
            count_m = re.search(r'"count"\s*:\s*(\d+)', m.group(0))
            count = int(count_m.group(1)) if count_m else None
            if name.lower() != keyword.lower():  # 完全一致のみ除外はしない
                brands.append({"name": name, "id": bid, "count": count})
    # 2) HTMLのブランドフィルタ要素から取得
    for el in soup.select('[class*="brand"] a, [class*="Brand"] a, [data-brand] a'):
        text = el.get_text(strip=True)
        m = re.match(r'(.+?)\s*[（(](\d+)[)）]', text)
        if m:
            name, count = m.group(1), int(m.group(2))
        else:
            name, count = text, None
        href = el.get('href', '')
        bid_m = re.search(r'brand(?:\[\])?=(\d+)', href)
        bid = bid_m.group(1) if bid_m else ""
        if name:
            brands.append({"name": name, "id": bid, "count": count})
    # 3) チェックボックス形式のブランドフィルタ
    for el in soup.select('input[name*="brand"]'):
        bid = el.get('value', '')
        label_el = el.find_next('label') or el.find_parent('label')
        if label_el:
            text = label_el.get_text(strip=True)
            m = re.match(r'(.+?)\s*[（(](\d+)[)）]', text)
            if m:
                name, count = m.group(1), int(m.group(2))
            else:
                name, count = text, None
            if name and bid:
                brands.append({"name": name, "id": bid, "count": count})
    # 重複排除（name基準）
    seen = set()
    unique = []
    for b in brands:
        key = b["name"].lower()
        if key not in seen:
            seen.add(key)
            unique.append(b)
    # キーワードでフィルタ（部分一致）
    kw = keyword.lower()
    filtered = [b for b in unique if kw in b["name"].lower()]
    # 件数降順ソート
    filtered.sort(key=lambda b: b["count"] or 0, reverse=True)
    return filtered

def _check_count_secondstreet(brand, category, scraper_api_key=""):
    """セカスト: 検索結果件数を概算取得（本体と同じフォールバック順）"""
    url = secondstreet_build_url(brand, category)
    method = ""
    debug = ""
    try:
        html = None
        # 0a) ScraperAPI render=true + wait（JS完全実行）
        if scraper_api_key:
            try:
                api_url = (f"https://api.scraperapi.com?api_key={scraper_api_key}"
                           f"&url={url}&render=true&wait=5000")
                res = requests.get(api_url, timeout=90)
                if res.status_code == 200:
                    html = res.text
                    method = "ScraperAPI(render)"
                    count, info = _parse_secondstreet_count(html, method)
                    if count > 0:
                        return count, url, info
                    debug += f"ScraperAPI(render): HTML取得済みだが件数0 ({info}); "
                    html = None  # 次の方式を試す
                else:
                    debug += f"ScraperAPI(render): HTTP {res.status_code}; "
            except Exception as e:
                debug += f"ScraperAPI(render): {e}; "
        # 0b) ScraperAPI render=false（本体スクレイパーと同じ設定）
        if not html and scraper_api_key:
            try:
                api_url = (f"https://api.scraperapi.com?api_key={scraper_api_key}"
                           f"&url={url}&render=false")
                res = requests.get(api_url, timeout=60)
                if res.status_code == 200:
                    html = res.text
                    method = "ScraperAPI(norender)"
                    count, info = _parse_secondstreet_count(html, method)
                    if count > 0:
                        return count, url, info
                    debug += f"ScraperAPI(norender): HTML取得済みだが件数0 ({info}); "
                    html = None
                else:
                    debug += f"ScraperAPI(norender): HTTP {res.status_code}; "
            except Exception as e:
                debug += f"ScraperAPI(norender): {e}; "
        if not scraper_api_key:
            debug += "ScraperAPI: キーなし; "
        # 1) Chrome（undetected-chromedriver）
        if not html:
            html = _fetch_with_chrome(url)
            if html:
                method = "Chrome"
            else:
                debug += "Chrome: 失敗; "
        # 2) curl_cffi（TLS指紋偽装）
        if not html:
            html = _fetch_with_curl(url)
            if html:
                method = "curl_cffi"
            else:
                debug += "curl_cffi: 失敗; "
        # 3) cloudscraper
        if not html and _scraper:
            try:
                res = _scraper.get(url, headers=HEADERS, timeout=15)
                if res.status_code == 200:
                    html = res.text
                    method = "cloudscraper"
                else:
                    debug += f"cloudscraper: HTTP {res.status_code}; "
            except Exception as e:
                debug += f"cloudscraper: {e}; "
        # 4) requests Session（Cookie保持で403回避）
        if not html:
            try:
                ss = _get_secondstreet_session()
                res = ss.get(url, timeout=30)
                if res.status_code == 200:
                    html = res.text
                    method = "requests"
                else:
                    debug += f"requests: HTTP {res.status_code}; "
            except Exception as e:
                debug += f"requests: {e}; "
        if not html:
            return None, url, f"全方式失敗 ({debug})"
        count, info = _parse_secondstreet_count(html, method)
        if count > 0:
            return count, url, info
        return 0, url, f"{debug}{info}"
    except Exception as e:
        return None, url, f"例外: {e}"

def check_product_counts(brand, category, selected_sites):
    """選択サイトの商品件数を一括チェック"""
    scraper_api_key = st.session_state.get("scraper_api_key", "")
    other_checkers = {
        "コメ兵": _check_count_komehyo,
        "RAGTAG": _check_count_ragtag,
        "トレファク": _check_count_trefac,
    }
    results = {}
    # セカスト以外は並列実行
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {}
        for site in selected_sites:
            if site in other_checkers:
                futures[executor.submit(other_checkers[site], brand, category)] = site
        for future in as_completed(futures):
            site = futures[future]
            try:
                count, url = future.result()
                results[site] = {"count": count, "url": url}
            except Exception:
                results[site] = {"count": None, "url": ""}
    # セカストはメインスレッドで実行（st.session_state対応 + デバッグ出力）
    if "セカスト" in selected_sites:
        try:
            count, url, debug_msg = _check_count_secondstreet(brand, category, scraper_api_key)
            results["セカスト"] = {"count": count, "url": url, "debug": debug_msg}
        except Exception as e:
            results["セカスト"] = {"count": None, "url": "", "debug": str(e)}
    return results

def to_excel(df):
    """DataFrameをExcelバイナリに変換（サイト列がある場合はシート分割）"""
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if "サイト" in df.columns:
                df.to_excel(writer, index=False, sheet_name='全サイト')
                for site_name in df["サイト"].unique():
                    site_df = df[df["サイト"] == site_name].drop(columns=["サイト"])
                    sheet_name = site_name[:31]
                    site_df.to_excel(writer, index=False, sheet_name=sheet_name)
            else:
                df.to_excel(writer, index=False, sheet_name='データ')
    except Exception:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='データ')
    return output.getvalue()
# ===== 品番→正式名称取得用関数（v2: ローカル優先 + Web検索補完） =====
_BRAND_JP_MAP = {
    'fendi': 'フェンディ', 'gucci': 'グッチ', 'prada': 'プラダ',
    'chanel': 'シャネル', 'hermes': 'エルメス', 'celine': 'セリーヌ',
    'loewe': 'ロエベ', 'louisvuitton': 'ルイヴィトン', 'louis vuitton': 'ルイヴィトン',
    'dior': 'ディオール', 'balenciaga': 'バレンシアガ',
    'bottegaveneta': 'ボッテガヴェネタ', 'bottega veneta': 'ボッテガヴェネタ',
    'saintlaurent': 'サンローラン', 'saint laurent': 'サンローラン',
    'burberry': 'バーバリー', 'givenchy': 'ジバンシィ',
    'miumiu': 'ミュウミュウ', 'miu miu': 'ミュウミュウ',
    'valentino': 'ヴァレンティノ', 'versace': 'ヴェルサーチ',
    'jimmychoo': 'ジミーチュウ', 'jimmy choo': 'ジミーチュウ',
    'chloe': 'クロエ', 'marni': 'マルニ', 'coach': 'コーチ',
    'michaelkors': 'マイケルコース', 'michael kors': 'マイケルコース',
    'todayful': 'トゥデイフル', 'isabelmarant': 'イザベルマラン',
}
_NOISE_WORDS_JA = [
    'ハンドバッグ', 'ショルダーバッグ', 'トートバッグ', 'ボストンバッグ',
    'クラッチバッグ', 'クロスボディバッグ', 'メッセンジャーバッグ', 'ボディバッグ',
    'ウエストバッグ', 'ミニバッグ', 'バックパック', 'リュックサック', 'リュック',
    'ポーチ', 'ウォレット', 'チェーンウォレット',
    '長財布', '二つ折り財布', '三つ折り財布', '財布', 'コインケース',
    'カードケース', 'キーケース', 'パスケース', '名刺入れ',
    'ベルト', 'スカーフ', 'マフラー', 'ストール', 'ネクタイ',
    'サンダル', 'スニーカー', 'パンプス', 'ブーツ', 'ローファー',
    'バッグ', '2WAY', '2Way', '2way', '3WAY', '3Way',
    '中古', '美品', '新品', '未使用', 'A品', 'B品', 'C品', 'S品', 'N品',
    'メンズ', 'レディース', 'ユニセックス',
    'レザー', 'キャンバス', 'ナイロン', 'スエード', 'デニム',
    'コットン', 'シルク', 'ウール', 'カシミヤ',
]
_NOISE_WORDS_EN = [
    'Handbag', 'Shoulder Bag', 'Tote Bag', 'Crossbody', 'Clutch',
    'Backpack', 'Wallet', 'Pouch', 'Satchel', 'Boston',
    'Bag', 'Tote', '2Way',
    'Pre-Owned', 'Authentic', 'Used', 'New',
    'Leather', 'Canvas', 'Nylon', 'Suede', 'Calfskin', 'Lambskin',
    'Black', 'Brown', 'White', 'Red', 'Blue', 'Green', 'Pink',
    'Beige', 'Grey', 'Gray', 'Navy', 'Gold', 'Silver', 'Cream',
    'Yellow', 'Orange', 'Purple', 'Ivory', 'Tan',
    'Ladies', 'Womens', "Women's", 'Mens', "Men's",
]
def _get_brand_variants(brand):
    """ブランド名のバリエーション一覧を取得"""
    brand_lower = brand.lower().strip().replace(" ", "")
    brand_with_space = brand.lower().strip()
    variants = {brand.strip(), brand.lower(), brand.upper(), brand_lower}
    jp_name = _BRAND_JP_MAP.get(brand_lower, "") or _BRAND_JP_MAP.get(brand_with_space, "")
    if jp_name:
        variants.add(jp_name)
    return variants
def _is_valid_hinban(val):
    """品番として有効かチェック。"""
    if not val or len(val) < 3:
        return False
    if re.fullmatch(r'\d{10,}', val):
        return False
    if re.fullmatch(r'\d{2}(?:AW|SS|FW|PF|PRE|RE)', val, re.IGNORECASE):
        return False
    if re.fullmatch(r'\d+\.?\d*\s*cm', val, re.IGNORECASE):
        return False
    if re.fullmatch(r'(?:SIZE|サイズ)\s*\d+', val, re.IGNORECASE):
        return False
    if re.fullmatch(r'[XSML]+', val, re.IGNORECASE):
        return False
    if re.fullmatch(r'FREE\s*(?:SIZE)?', val, re.IGNORECASE):
        return False
    if re.fullmatch(r'(?:SV|K|PT)\d+', val, re.IGNORECASE):
        return False
    if ' ' in val and len(val) > 20:
        return False
    return True
def extract_unique_hinban(df):
    """スクレイピング結果から有効な品番/型番を抽出・出品数順でソート。"""
    from collections import Counter
    freq = Counter()
    if "品番" in df.columns:
        for val in df["品番"].dropna():
            val_str = str(val).strip()
            if _is_valid_hinban(val_str):
                freq[val_str] += 1
    if "型番" in df.columns:
        for val in df["型番"].dropna():
            val_str = str(val).strip()
            if val_str and len(val_str) >= 3:
                for part in val_str.split(" / "):
                    part = part.strip()
                    if _is_valid_hinban(part):
                        freq[part] += 1
    return [h for h, _ in sorted(freq.items(), key=lambda x: (-x[1], x[0][::-1]))]
def _clean_product_name(name, brand_variants, hinban):
    """商品名からブランド名・ノイズワード・品番を除去して正式名称部分を抽出"""
    cleaned = name.strip()
    if not cleaned:
        return ""
    for bv in sorted(brand_variants, key=len, reverse=True):
        cleaned = re.sub(re.escape(bv), ' ', cleaned, flags=re.IGNORECASE)
    if hinban:
        cleaned = re.sub(re.escape(hinban), ' ', cleaned, flags=re.IGNORECASE)
    for nw in _NOISE_WORDS_JA:
        cleaned = cleaned.replace(nw, ' ')
    for nw in _NOISE_WORDS_EN:
        cleaned = re.sub(r'\b' + re.escape(nw) + r'\b', ' ', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\b\d{2}(?:AW|SS|FW|PF|PRE|RE)\b', ' ', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'[￥¥][\d,]+', ' ', cleaned)
    cleaned = re.sub(r'\$[\d,.]+', ' ', cleaned)
    cleaned = re.sub(r'\d+円', ' ', cleaned)
    cleaned = re.sub(r'[()（）\[\]【】「」/\\|,、。・\-–—:：;；#＃]+', ' ', cleaned)
    cleaned = re.sub(r'\.{2,}', ' ', cleaned)
    cleaned = re.sub(r'\b\d+\b', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    if len(cleaned) < 2 or re.fullmatch(r'[\d\s]+', cleaned):
        return ""
    if re.fullmatch(r'\d{2}(?:AW|SS|FW|PF|PRE|RE)', cleaned, re.IGNORECASE):
        return ""
    _CATEGORY_ONLY = {
        'コート', 'ジャケット', 'パンツ', 'スカート', 'ワンピース', 'ブラウス', 'シャツ',
        'ニット', 'セーター', 'カーディガン', 'ベスト', 'Tシャツ', 'トップス', 'ボトムス',
        'デニム', 'ドレス', 'チュニック', 'キャミソール', 'タンクトップ', 'スウェット',
        'パーカー', 'フーディー', 'オールインワン', 'サロペット', 'ロンパース',
        'レギンス', 'ストレートパンツ', 'ワイドパンツ', 'テーパードパンツ',
        'アウター', 'トレーナー', 'カットソー', 'プルオーバー',
        'ボトム', '長袖シャツ', 'セットアップ', 'ローカット', 'ハイカット',
        'キャミワンピース', 'ノースリーブワンピース',
        'ニット セーター 薄手', 'カーディガン 薄手',
        'ダウンジャケット', 'ダウンコート', 'ロングコート', 'チェスターコート',
        'テーラードジャケット', 'ライダースジャケット', 'ナイロンジャケット',
        'マウンテンパーカー', 'モッズコート', 'ピーコート', 'トレンチコート',
        'ステンカラーコート', 'ダッフルコート', 'ボアジャケット',
        'ロングスカート', 'ミニスカート', 'ミディスカート', 'フレアスカート',
        'ショートパンツ', 'クロップドパンツ', 'スキニーパンツ',
        'ロングワンピース', 'シャツワンピース', 'ニットワンピース',
        'スニーカー', 'サンダル', 'パンプス', 'ブーツ', 'ローファー', 'フラットシューズ',
    }
    if cleaned in _CATEGORY_ONLY:
        return ""
    return cleaned
def extract_names_from_local_data(df, brand):
    """Phase 1: スクレイピングデータから品番→正式名称マッピングを作成"""
    from collections import Counter
    brand_variants = _get_brand_variants(brand)
    hinban_to_data = {}
    for _, row in df.iterrows():
        hinbans = []
        if "品番" in df.columns and pd.notna(row.get("品番")):
            h = str(row["品番"]).strip()
            if h and not re.fullmatch(r'\d{10,}', h):
                hinbans.append(h)
        if "型番" in df.columns and pd.notna(row.get("型番")):
            for part in str(row["型番"]).split(" / "):
                part = part.strip()
                if part and len(part) >= 3 and not re.fullmatch(r'\d{10,}', part):
                    hinbans.append(part)
        tsusho = ""
        if "通称" in df.columns and pd.notna(row.get("通称")):
            tsusho = str(row["通称"]).strip()
        shohin_name = ""
        if "商品名" in df.columns and pd.notna(row.get("商品名")):
            shohin_name = str(row["商品名"]).strip()
        for h in hinbans:
            if h not in hinban_to_data:
                hinban_to_data[h] = {"通称": [], "商品名": []}
            if tsusho:
                hinban_to_data[h]["通称"].append(tsusho)
            if shohin_name:
                hinban_to_data[h]["商品名"].append(shohin_name)
    results = {}
    for hinban, data in hinban_to_data.items():
        if data["通称"]:
            counter = Counter(data["通称"])
            best = counter.most_common(1)[0][0]
            results[hinban] = best
            continue
        names = data["商品名"]
        if not names:
            results[hinban] = ""
            continue
        cleaned_names = []
        for name in names:
            cleaned = _clean_product_name(name, brand_variants, hinban)
            if cleaned:
                cleaned_names.append(cleaned)
        if not cleaned_names:
            results[hinban] = ""
            continue
        normalized_map = {}
        for c in cleaned_names:
            key = c.lower().strip()
            if key not in normalized_map:
                normalized_map[key] = c
        counter = Counter(c.lower().strip() for c in cleaned_names)
        best_key = counter.most_common(1)[0][0]
        best_name = normalized_map[best_key]
        if len(best_name) > 60:
            best_name = best_name[:60].rsplit(' ', 1)[0]
        results[hinban] = best_name
    return results
def _wear_fetch_brand_catalog(brand, progress_callback=None):
    """WEARからブランドの全商品カタログを取得し、brand_no→名称の辞書を返す。"""
    session = requests.Session()
    api_headers = {
        'User-Agent': HEADERS['User-Agent'],
        'Accept': 'application/json',
        'Origin': 'https://wear.jp',
        'Referer': 'https://wear.jp/',
    }
    try:
        res = session.get(
            f"https://api.wear2.jp/connect/v1/brands/search?keyword={requests.utils.quote(brand)}",
            headers=api_headers, timeout=10,
        )
        if res.status_code != 200:
            return {}
        data = res.json()
        pm = data.get('perfect_match_brand')
        if not pm:
            brands = data.get('brands', [])
            if not brands:
                return {}
            pm = brands[0].get('brand', {})
        brand_id = pm.get('id')
        items_count = pm.get('brand_aggregate', {}).get('items_count', 0)
        if not brand_id:
            return {}
    except Exception:
        return {}
    if progress_callback:
        progress_callback(0, 0, f"👗 WEAR: {brand.upper()} カタログ取得中（{items_count}件）...")
    catalog = {}
    page_size = 60
    max_pages = (items_count // page_size) + 2
    max_pages = min(max_pages, 200)
    for page_num in range(1, max_pages + 1):
        try:
            url = f"https://api.wear2.jp/connect/v2/items/search?brand_id={brand_id}&pageno={page_num}&count={page_size}"
            res = session.get(url, headers=api_headers, timeout=15)
            if res.status_code != 200:
                break
            items = res.json().get('items', [])
            if not items:
                break
            for item in items:
                bn = str(item.get('brand_no', '')).strip()
                raw_name = str(item.get('name', '')).strip()
                if not bn or not raw_name:
                    continue
                bn_keys = [b.strip() for b in bn.split('/') if b.strip()]
                cleaned = re.sub(r'^[\[【][^\]】]*[\]】]\s*', '', raw_name)
                cleaned = re.sub(r'/[春夏秋冬]服$', '', cleaned)
                cleaned = re.sub(r'/\d{7,}$', '', cleaned)
                if '/' in cleaned:
                    parts = [p.strip() for p in cleaned.split('/') if p.strip()]
                    if len(parts) >= 2:
                        jp_parts = [p for p in parts if not re.fullmatch(r'[A-Za-z0-9\s\-\.\'\",()]+', p)]
                        en_parts = [p for p in parts if re.fullmatch(r'[A-Za-z0-9\s\-\.\'\",()]+', p)]
                        if jp_parts:
                            cleaned = jp_parts[0]
                        elif en_parts:
                            cleaned = en_parts[0]
                        else:
                            cleaned = parts[0]
                    else:
                        cleaned = parts[0]
                cleaned = re.sub(r'^TODAYFUL\s*', '', cleaned, flags=re.IGNORECASE).strip()
                if cleaned and len(cleaned) >= 2:
                    for key in bn_keys:
                        if key not in catalog:
                            catalog[key] = cleaned
            if progress_callback and page_num % 10 == 0:
                progress_callback(0, 0, f"👗 WEAR: {len(catalog)}件取得済み（page {page_num}）...")
        except Exception:
            continue
        time.sleep(0.15)
    return catalog
def _wear_search_batch(hinban_list, brand, progress_callback=None, start_idx=0, total=0):
    """Phase 2a: WEAR API（api.wear2.jp）で品番の正式名称を取得。"""
    brand_key = brand.lower().strip()
    try:
        cache = st.session_state.get('wear_catalog_cache', {})
        if brand_key in cache:
            catalog = cache[brand_key]
            if progress_callback:
                progress_callback(start_idx, total, f"👗 WEAR: キャッシュ済みカタログ{len(catalog)}件から品番マッチング中...")
        else:
            catalog = _wear_fetch_brand_catalog(brand, progress_callback=progress_callback)
            if catalog:
                st.session_state.wear_catalog_cache[brand_key] = catalog
    except Exception:
        catalog = _wear_fetch_brand_catalog(brand, progress_callback=progress_callback)
    if not catalog:
        return {}
    if progress_callback:
        progress_callback(start_idx, total, f"👗 WEAR: カタログ{len(catalog)}件から品番マッチング中...")
    results = {}
    for hinban in hinban_list:
        name = catalog.get(hinban, "")
        if name:
            results[hinban] = name
    return results
def _web_search_batch(hinban_list, brand, progress_callback=None, start_idx=0, total=0):
    """Phase 2b: Web検索（DDG HTML → Brave）で品番の正式名称を補完"""
    brand_variants = _get_brand_variants(brand)
    results = {}
    remaining = list(hinban_list)
    ddg_session = requests.Session()
    for i, hinban in enumerate(remaining[:]):
        if hinban in results:
            continue
        if progress_callback:
            progress_callback(start_idx + i, total, f"🌐 [{start_idx + i + 1}/{total}] {hinban} をWeb検索中...")
        q = f"{brand} {hinban}"
        url = f"https://html.duckduckgo.com/html/?q={requests.utils.quote(q)}"
        try:
            res = ddg_session.get(url, headers=HEADERS, timeout=15)
            if res.status_code == 200:
                soup = BeautifulSoup(res.text, 'html.parser')
                titles = []
                for r in soup.select('.result'):
                    title_el = r.select_one('.result__title')
                    if title_el:
                        titles.append(title_el.get_text(strip=True))
                if titles:
                    name = _extract_name_from_search_titles(titles[:8], hinban, brand_variants)
                    if name:
                        results[hinban] = name
                        remaining.remove(hinban)
            elif res.status_code in (202, 403, 429):
                break
        except Exception:
            break
        time.sleep(random.uniform(2.0, 3.5))
    if not remaining:
        return results
    brave_session = requests.Session()
    for i, hinban in enumerate([h for h in remaining if h not in results]):
        if progress_callback:
            idx = start_idx + len(hinban_list) - len(remaining) + i
            progress_callback(idx, total, f"🌐 [{idx + 1}/{total}] {hinban} をBrave検索中...")
        q = f"{brand} {hinban}"
        url = f"https://search.brave.com/search?q={requests.utils.quote(q)}"
        try:
            res = brave_session.get(url, headers=HEADERS, timeout=10)
            if res.status_code == 200:
                soup = BeautifulSoup(res.text, 'html.parser')
                titles = [el.get_text(strip=True) for el in soup.select('#results .snippet .title') if el.get_text(strip=True)]
                if titles:
                    name = _extract_name_from_search_titles(titles[:8], hinban, brand_variants)
                    if name:
                        results[hinban] = name
                        if hinban in remaining:
                            remaining.remove(hinban)
            elif res.status_code in (403, 429):
                break
        except Exception:
            break
        time.sleep(random.uniform(2.0, 3.0))
    return results
def _extract_name_from_search_titles(titles, hinban, brand_variants):
    """Web検索タイトル群から正式名称を抽出"""
    from collections import Counter
    site_patterns = [
        r'\s*[\-–—|]\s*(eBay|Grailed|Poshmark|Amazon|Vestiaire|Yoogi.*?Closet|eLADY|Rebag|'
        r'Fashionphile|TheRealReal|Mercari|Depop|StockX|Farfetch|Jomashop|Nordstrom|'
        r'Ssense|Mytheresa|Joli\s*Closet|コメ兵|RAGTAG|ラグタグ|トレファク|'
        r'セカンドストリート|メルカリ|楽天|BUYMA|Yahoo).*$',
    ]
    candidates = []
    for title in titles:
        cleaned = title
        for pat in site_patterns:
            cleaned = re.sub(pat, '', cleaned, flags=re.IGNORECASE)
        cleaned = _clean_product_name(cleaned, brand_variants, hinban)
        if cleaned and len(cleaned) >= 2:
            candidates.append(cleaned)
    if not candidates:
        return ""
    normalized_map = {}
    for c in candidates:
        key = c.lower().strip()
        if key not in normalized_map:
            normalized_map[key] = c
    counter = Counter(c.lower().strip() for c in candidates)
    best_key = counter.most_common(1)[0][0]
    best_name = normalized_map[best_key]
    if len(best_name) > 60:
        best_name = best_name[:60].rsplit(' ', 1)[0]
    return best_name
def lookup_official_names(df, brand, max_lookup=50, web_search=True, wear_search=True, progress_callback=None):
    """品番→正式名称を3フェーズで一括取得"""
    unique_hinban = extract_unique_hinban(df)
    target = unique_hinban[:max_lookup]
    total = len(target)
    wear_results = {}
    if wear_search:
        if progress_callback:
            progress_callback(0, total, "👗 Phase 1: WEARカタログから正式名称を取得中...")
        wear_results = _wear_search_batch(
            target, brand,
            progress_callback=progress_callback,
            start_idx=0,
            total=total,
        )
    wear_count = sum(1 for h in target if wear_results.get(h, ""))
    if progress_callback:
        progress_callback(wear_count, total, f"👗 Phase 1完了: {wear_count}/{total}件をWEARで特定")
    if progress_callback:
        progress_callback(wear_count, total, "📊 Phase 2: ローカルデータから正式名称を抽出中...")
    local_results = extract_names_from_local_data(df, brand)
    resolved_count = sum(1 for h in target if wear_results.get(h, "") or local_results.get(h, ""))
    if progress_callback:
        progress_callback(resolved_count, total, f"📊 Phase 2完了: {resolved_count}/{total}件を特定")
    still_unresolved = [h for h in target if not wear_results.get(h, "") and not local_results.get(h, "")]
    web_results = {}
    if web_search and still_unresolved:
        if progress_callback:
            progress_callback(resolved_count, total, f"🌐 Phase 3: 残り{len(still_unresolved)}件をWeb検索中...")
        web_results = _web_search_batch(
            still_unresolved, brand,
            progress_callback=progress_callback,
            start_idx=resolved_count,
            total=total,
        )
    mapping_results = []
    for hinban in target:
        wear_name = wear_results.get(hinban, "")
        local_name = local_results.get(hinban, "")
        web_name = web_results.get(hinban, "")
        final_name = wear_name or local_name or web_name
        source = "WEAR" if wear_name else ("ローカル" if local_name else ("Web検索" if web_name else ""))
        mapping_results.append({
            "品番": hinban,
            "正式名称": final_name,
            "ソース": source,
        })
    if progress_callback:
        found = sum(1 for r in mapping_results if r["正式名称"])
        progress_callback(total, total, f"✅ 完了！{found}/{total}件の正式名称を特定")
    return pd.DataFrame(mapping_results)
# ===== KDE価格分析 =====
def _kde_peak_price(prices):
    """価格リストからKDE（カーネル密度推定）でピーク価格を算出。"""
    prices = [p for p in prices if p and p > 0]
    if not prices:
        return 0
    if len(prices) == 1:
        return prices[0]
    arr = np.array(prices, dtype=float)
    n = len(arr)
    std = arr.std()
    if std == 0:
        return float(np.median(arr))
    iqr = float(np.percentile(arr, 75) - np.percentile(arr, 25))
    bandwidth = 0.9 * min(std, iqr / 1.34 if iqr > 0 else std) * n ** (-0.2)
    bandwidth = max(bandwidth, 1000)
    x_min = arr.min() - 3 * bandwidth
    x_max = arr.max() + 3 * bandwidth
    x = np.linspace(x_min, x_max, 500)
    density = np.zeros_like(x)
    for p in arr:
        density += np.exp(-0.5 * ((x - p) / bandwidth) ** 2)
    peak_idx = int(np.argmax(density))
    return round(float(x[peak_idx]))
def enrich_mapping_with_kde(df, mapping_df, min_price=10000):
    """マッピングテーブルにKDE価格分析を追加し、ピーク価格が閾値以下の品番を除外。"""
    hinban_prices = {}
    for _, row in df.iterrows():
        price = row.get("価格")
        if pd.isna(price) or price is None or float(price) <= 0:
            continue
        hinbans = []
        if "品番" in df.columns and pd.notna(row.get("品番")):
            h = str(row["品番"]).strip()
            if h and not re.fullmatch(r'\d{10,}', h):
                hinbans.append(h)
        if "型番" in df.columns and pd.notna(row.get("型番")):
            for part in str(row["型番"]).split(" / "):
                part = part.strip()
                if part and len(part) >= 3 and not re.fullmatch(r'\d{10,}', part):
                    hinbans.append(part)
        for h in hinbans:
            if h not in hinban_prices:
                hinban_prices[h] = []
            hinban_prices[h].append(float(price))
    counts = []
    kde_prices = []
    price_ranges = []
    for _, row in mapping_df.iterrows():
        hinban = row["品番"]
        prices = hinban_prices.get(hinban, [])
        counts.append(len(prices))
        if prices:
            peak = _kde_peak_price(prices)
            kde_prices.append(peak)
            if len(prices) > 1:
                price_ranges.append(f"¥{min(prices):,.0f}〜¥{max(prices):,.0f}")
            else:
                price_ranges.append(f"¥{prices[0]:,.0f}")
        else:
            kde_prices.append(0)
            price_ranges.append("")
    enriched = mapping_df.copy()
    enriched["件数"] = counts
    enriched["KDE代表価格"] = kde_prices
    enriched["価格レンジ"] = price_ranges
    filtered = enriched[
        (enriched["正式名称"] != "") &
        ((enriched["KDE代表価格"] > min_price) | (enriched["KDE代表価格"] == 0))
    ].copy()
    return enriched, filtered
# タイトル
st.title("👜 ブランドECスクレイパー")
st.caption("ECサイトとブランドを選んで商品データを取得")
# サイドバー
with st.sidebar:
    st.markdown("### 🏬 ECサイト選択")
    all_site_names = list(EC_SITES.keys())
    selected_sites = st.multiselect(
        "スクレイピング先（複数選択可）",
        options=all_site_names,
        default=all_site_names,
        format_func=lambda x: f"{EC_SITES[x]['icon']} {x}",
    )
    st.session_state.selected_sites = selected_sites
    st.divider()
    st.markdown("### 🔧 検索条件")
    brand_input = st.text_input(
        "ブランド名（英語）",
        value="fendi",
        help="スペースなしの英語で入力"
    )
    if len(selected_sites) == 1:
        selected_ec = selected_sites[0]
        if selected_ec == "コメ兵":
            st.caption("📝 例: fendi, gucci, prada, chanel, hermes, celine, loewe, louisvuitton")
            categories = KOMEHYO_CATEGORIES
        elif selected_ec == "RAGTAG":
            st.caption("📝 例: FENDI, GUCCI, PRADA, CHANEL, HERMES, CELINE, LOEWE（大文字推奨）")
            categories = RAGTAG_CATEGORIES
        elif selected_ec == "セカスト":
            st.caption("📝 例: fendi, gucci, prada, chanel, hermes, celine, loewe")
            # ブランドサジェスト機能（全候補一括検索）
            suggest_btn = st.button("🔍 ブランド候補を検索", key="brand_suggest_btn",
                                    use_container_width=True,
                                    help="入力中のブランド名でセカストのブランド候補を取得し、全候補を一括スクレイピングできます")
            if suggest_btn and brand_input and len(brand_input) >= 2:
                scraper_key = st.session_state.get("scraper_api_key", "")
                with st.spinner("🔍 ブランド候補を検索中..."):
                    results = _suggest_brands_secondstreet(brand_input, scraper_key)
                st.session_state.brand_suggest_results = results
                st.session_state.brand_suggest_keyword = brand_input
            if (st.session_state.brand_suggest_results
                    and st.session_state.brand_suggest_keyword.lower() in brand_input.lower()):
                suggestions = st.session_state.brand_suggest_results
                st.markdown(f"**🏷️ ブランド候補（{len(suggestions)}件）**")
                for b in suggestions:
                    count_str = f"（{b['count']}件）" if b.get('count') else ""
                    st.caption(f"ブランド： {b['name']} {count_str}")
                st.info(f"🔍 スクレイピング開始で上記 {len(suggestions)} ブランドを全て一括検索します")
            categories = SECONDSTREET_CATEGORIES
        else:
            st.caption("📝 例: fendi, gucci, prada, chanel, hermes, celine, loewe")
            categories = TREFAC_CATEGORIES
        category_name = st.selectbox(
            "カテゴリ",
            options=list(categories.keys()),
            index=0
        )
        category = categories[category_name]
    else:
        st.caption("📝 例: fendi, gucci, prada, chanel, hermes, celine, loewe")
        if len(selected_sites) >= 2:
            st.info("📂 複数サイト選択時はカテゴリ: 全カテゴリ で実行されます")
        category = ""
    all_pages = st.checkbox("全件取得", value=False)
    if all_pages:
        max_pages = 999
    else:
        if len(selected_sites) == 1:
            selected_ec = selected_sites[0]
            if selected_ec == "コメ兵":
                max_pages = st.slider("取得ページ数", 1, 200, 10, help="1ページ約50件（200p=約10,000件）")
            elif selected_ec == "RAGTAG":
                max_pages = st.slider("取得ページ数", 1, 200, 5, help="1ページ約100件（200p=約20,000件）")
            elif selected_ec == "セカスト":
                max_pages = st.slider("取得ページ数", 1, 200, 5, help="1ページ約60件（200p=約12,000件）")
            else:
                max_pages = st.slider("取得ページ数", 1, 200, 5, help="1ページ約90件（200p=約18,000件）")
        else:
            max_pages = st.slider("取得ページ数", 1, 200, 10, help="各サイトごとのページ数")
    st.divider()
    # ScraperAPI設定（セカスト用・Streamlit Cloud対応）
    # st.secrets に SCRAPER_API_KEY があれば自動で使う
    secrets_key = st.secrets.get("SCRAPER_API_KEY", "") if hasattr(st, "secrets") else ""
    if secrets_key and "scraper_api_key" not in st.session_state:
        st.session_state["scraper_api_key"] = secrets_key
    with st.expander("🔑 ScraperAPI設定（セカスト用）", expanded=False):
        st.caption("Streamlit Cloudでセカストが403になる場合に使用")
        st.caption("[無料アカウント作成（5000回/月）](https://www.scraperapi.com/signup)")
        default_key = st.session_state.get("scraper_api_key", "")
        api_key = st.text_input("APIキー", value=default_key, type="password", key="scraper_api_key_input")
        if api_key:
            st.session_state["scraper_api_key"] = api_key
            st.success("APIキー設定済み")
        else:
            st.session_state["scraper_api_key"] = ""
    st.divider()
    # multiselectタグの白背景をJS MutationObserverで強制適用
    # （Styletronが後からprimaryColorを注入するためCSSだけでは上書き不可）
    import streamlit.components.v1 as _components
    _components.html("""
    <script>
    (function() {
        var doc = window.parent.document;
        function fixTags() {
            doc.querySelectorAll('[data-baseweb="tag"]').forEach(function(tag) {
                tag.style.setProperty('background-color', '#FFFFFF', 'important');
                tag.style.setProperty('background', '#FFFFFF', 'important');
                tag.style.setProperty('color', '#31333F', 'important');
                tag.style.setProperty('border', '1px solid #D0D0D0', 'important');
                tag.style.setProperty('border-radius', '6px', 'important');
            });
        }
        fixTags();
        new MutationObserver(fixTags).observe(doc.body, {
            childList: true, subtree: true, attributes: true
        });
    })();
    </script>
    """, height=0)
    count_button = st.button("📊 件数チェック", use_container_width=True, key="count_check_btn")
    scrape_button = st.button("🔍 スクレイピング開始", type="primary", use_container_width=True)
    if st.session_state.scraping_done:
        if st.button("🗑️ 結果をクリア", use_container_width=True):
            st.session_state.results_df = None
            st.session_state.scraping_done = False
            st.session_state.brand_name = ""
            st.session_state.hinban_mapping_df = None
            st.session_state.hinban_lookup_done = False
            st.session_state.hinban_merged_df = None
            st.session_state.wear_catalog_cache = {}
            _clear_results_backup()
            st.rerun()
# 件数チェック処理
if count_button:
    if not brand_input:
        st.warning("⚠️ ブランド名を入力してください")
    elif not selected_sites:
        st.warning("⚠️ サイトを1つ以上選択してください")
    else:
        # セカスト用サジェスト候補
        _ss_suggest = []
        if (st.session_state.brand_suggest_results
                and st.session_state.brand_suggest_keyword.lower() in brand_input.lower()):
            _ss_suggest = [b["name"] for b in st.session_state.brand_suggest_results]
        with st.spinner("📊 件数を確認中..."):
            _cat = category if len(selected_sites) == 1 else ""
            count_results = check_product_counts(brand_input, _cat, selected_sites)
        total = 0
        for site_name in selected_sites:
            if site_name in count_results:
                info = count_results[site_name]
                icon = EC_SITES[site_name]["icon"]
                debug_msg = info.get("debug", "")
                if info["count"] is not None and info["count"] > 0:
                    total += info["count"]
                    st.success(f"{icon} **{site_name}**: 約 **{info['count']:,}件**")
                    if debug_msg:
                        st.caption(f"🔧 {debug_msg}")
                elif info["count"] == 0:
                    st.error(f"{icon} **{site_name}**: 約 **0件**")
                    if debug_msg:
                        st.warning(f"🔧 デバッグ: {debug_msg}")
                else:
                    st.error(f"{icon} **{site_name}**: 取得失敗")
                    if debug_msg:
                        st.warning(f"🔧 デバッグ: {debug_msg}")
        # セカスト＋サジェスト候補がある場合、各ブランド個別の件数も表示
        if "セカスト" in selected_sites and len(_ss_suggest) > 1:
            scraper_key = st.session_state.get("scraper_api_key", "")
            st.markdown("---")
            st.markdown("**🔴 セカスト ブランド別内訳**")
            brand_total = 0
            for bname in _ss_suggest:
                with st.spinner(f"🔴 {bname} の件数を確認中..."):
                    cnt, burl, dbg = _check_count_secondstreet(bname, _cat, scraper_key)
                if cnt and cnt > 0:
                    brand_total += cnt
                    st.caption(f"  {bname}: 約 **{cnt:,}件**")
                else:
                    st.caption(f"  {bname}: 0件")
            if brand_total > 0:
                st.success(f"🔴 **セカスト合計（全ブランド）**: 約 **{brand_total:,}件**")
        if total > 0:
            st.success(f"📦 合計: 約 **{total:,}件**（概算）")
# メイン処理
if scrape_button:
    if not brand_input:
        st.warning("⚠️ ブランド名を入力してください")
    elif not selected_sites:
        st.warning("⚠️ サイトを1つ以上選択してください")
    elif len(selected_sites) == 1:
        # ===== 単一サイト =====
        selected_ec = selected_sites[0]
        progress_bar = st.progress(0)
        status_text = st.empty()
        if selected_ec == "コメ兵":
            base_url = komehyo_build_url(brand_input, category)
            get_urls_func = komehyo_get_product_urls
            get_detail_func = komehyo_get_product_detail
        elif selected_ec == "RAGTAG":
            base_url = ragtag_build_url(brand_input, category)
            get_urls_func = ragtag_get_product_urls
            get_detail_func = ragtag_get_product_detail
        elif selected_ec == "セカスト":
            # ブランドサジェスト候補がある場合は全候補を一括検索
            _ss_brands = []
            if (st.session_state.brand_suggest_results
                    and st.session_state.brand_suggest_keyword.lower() in brand_input.lower()):
                _ss_brands = [b["name"] for b in st.session_state.brand_suggest_results]
            if len(_ss_brands) > 1:
                # 複数ブランド一括検索
                all_results = []
                total_brands = len(_ss_brands)
                overall_progress = st.progress(0)
                def _ss_brand_save_cb(cache_dict, _bi=brand_input):
                    _rescued = list(cache_dict.values())
                    _interim = all_results + _rescued
                    if _interim:
                        _df = pd.DataFrame(_interim)
                        _cols = ["ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                        _df = _df[[c for c in _cols if c in _df.columns]]
                        st.session_state.results_df = _df
                        st.session_state.brand_name = _bi
                        st.session_state.scraping_done = True
                        _save_results_backup(_df, _bi)
                for bi, bname in enumerate(_ss_brands):
                    status_text.text(f"🔴 [{bi+1}/{total_brands}] ブランド「{bname}」を検索中...")
                    overall_progress.progress(bi / total_brands)
                    burl = secondstreet_build_url(bname, category)
                    purls = secondstreet_get_product_urls(burl, max_pages, lambda msg: status_text.text(f"🔴 {bname}: {msg}"), save_callback=_ss_brand_save_cb)
                    if purls:
                        st.info(f"🔴 **{bname}**: {len(purls)}件発見")
                        brand_results = [secondstreet_get_product_detail(u) for u in purls]
                        brand_results = [r for r in brand_results if r]
                        all_results.extend(brand_results)
                overall_progress.progress(1.0)
                status_text.text("✅ 完了！")
                progress_bar.progress(1.0)
                if all_results:
                    df = pd.DataFrame(all_results)
                    columns = ["ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                    df = df[[c for c in columns if c in df.columns]]
                    st.session_state.results_df = df
                    st.session_state.brand_name = brand_input
                    st.session_state.scraping_done = True
                    _save_results_backup(df, brand_input)
                    st.rerun()
                else:
                    st.warning("⚠️ どのブランドからも商品を取得できませんでした")
                st.stop()
            base_url = secondstreet_build_url(brand_input, category)
            get_urls_func = secondstreet_get_product_urls
            get_detail_func = secondstreet_get_product_detail
        else:
            base_url = trefac_build_url(brand_input, category)
            get_urls_func = trefac_get_product_urls
            get_detail_func = trefac_get_product_detail
        status_text.text(f"🔗 URL: {base_url}")
        def update_status(msg):
            status_text.text(msg)
        _need_rerun = False
        results = []
        # セカスト: 10ページごとに中間保存するコールバック
        _single_save_cb = None
        if selected_ec == "セカスト":
            def _single_save_cb(cache_dict, _bi=brand_input):
                _rescued = list(cache_dict.values())
                _df = pd.DataFrame(_rescued)
                _cols = ["ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                _df = _df[[c for c in _cols if c in _df.columns]]
                st.session_state.results_df = _df
                st.session_state.brand_name = _bi
                st.session_state.scraping_done = True
                _save_results_backup(_df, _bi)
        try:
            product_urls = get_urls_func(base_url, max_pages, update_status, **({'save_callback': _single_save_cb} if _single_save_cb else {}))
            if not product_urls:
                st.error("❌ 商品が見つかりませんでした。ブランド名を確認してください。")
                st.info(f"試したURL: {base_url}")
            else:
                st.info(f"📦 {len(product_urls)}件の商品を発見")
                if selected_ec == "セカスト":
                    status_text.text("📦 データを整理中...")
                    results = [get_detail_func(url) for url in product_urls]
                    results = [r for r in results if r]
                else:
                    # コメ兵・RAGTAG・トレファク全て並列処理
                    status_text.text("🚀 並列処理で詳細取得中...")
                    def update_progress(completed, total):
                        progress_bar.progress(completed / total)
                        status_text.text(f"🚀 [{completed}/{total}] 並列取得中...")
                    results = get_products_parallel(product_urls, get_detail_func, max_workers=10, progress_callback=update_progress)
                status_text.text("✅ 完了！")
                progress_bar.progress(1.0)
                if results:
                    df = pd.DataFrame(results)
                    columns = ["ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                    df = df[[c for c in columns if c in df.columns]]
                    st.session_state.results_df = df
                    st.session_state.brand_name = brand_input
                    st.session_state.scraping_done = True
                    _save_results_backup(df, brand_input)
                    _need_rerun = True
                else:
                    st.warning("⚠️ 商品詳細の取得に失敗しました")
        except Exception as e:
            st.error(f"⚠️ スクレイピング中にエラーが発生しました: {e}")
            # セカストはキャッシュから部分データを救出
            if not results and _secondstreet_cache:
                results = list(_secondstreet_cache.values())
            # エラー発生時でも途中結果があれば保存
            if results:
                st.warning(f"⚠️ エラー発生前に取得済みの {len(results)}件 を保存します")
                df = pd.DataFrame(results)
                columns = ["ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                df = df[[c for c in columns if c in df.columns]]
                st.session_state.results_df = df
                st.session_state.brand_name = brand_input
                st.session_state.scraping_done = True
                _save_results_backup(df, brand_input)
                _need_rerun = True
        if _need_rerun:
            st.rerun()
    else:
        # ===== 複数サイト =====
        _cat = category
        all_site_configs = {
            "コメ兵": {
                "icon": "🏪",
                "build_url": lambda b, c=_cat: komehyo_build_url(b, c),
                "get_urls": komehyo_get_product_urls,
                "get_detail": komehyo_get_product_detail,
                "parallel": True,
            },
            "RAGTAG": {
                "icon": "👔",
                "build_url": lambda b, c=_cat: ragtag_build_url(b, c),
                "get_urls": ragtag_get_product_urls,
                "get_detail": ragtag_get_product_detail,
                "parallel": True,
            },
            "トレファク": {
                "icon": "🏷️",
                "build_url": lambda b, c=_cat: trefac_build_url(b, c),
                "get_urls": trefac_get_product_urls,
                "get_detail": trefac_get_product_detail,
                "parallel": True,
            },
            "セカスト": {
                "icon": "🔴",
                "build_url": lambda b, c=_cat: secondstreet_build_url(b, c),
                "get_urls": secondstreet_get_product_urls,
                "get_detail": secondstreet_get_product_detail,
                "parallel": False,
                "cached": True,
            },
        }
        targets = [(name, all_site_configs[name]) for name in selected_sites if name in all_site_configs]
        total_sites = len(targets)
        all_results = []
        overall_progress = st.progress(0)
        overall_status = st.empty()
        # セカスト用ブランドサジェスト候補
        _ss_suggest_brands = []
        if (st.session_state.brand_suggest_results
                and st.session_state.brand_suggest_keyword.lower() in brand_input.lower()):
            _ss_suggest_brands = [b["name"] for b in st.session_state.brand_suggest_results]
        for site_idx, (site_name, config) in enumerate(targets):
            try:
                site_icon = config["icon"]
                overall_status.text(f"{site_icon} [{site_idx+1}/{total_sites}] {site_name} をスクレイピング中...")
                overall_progress.progress(site_idx / total_sites)
                site_progress = st.progress(0)
                site_status = st.empty()
                # セカスト＋サジェスト候補がある場合は全候補を検索
                if site_name == "セカスト" and len(_ss_suggest_brands) > 1:
                    def _ms_ss_save_cb(cache_dict, _site=site_name, _bi=brand_input):
                        _rescued = list(cache_dict.values())
                        for _r in _rescued:
                            _r.setdefault("サイト", _site)
                        _all = all_results + _rescued
                        if _all:
                            _df = pd.DataFrame(_all)
                            _cols = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                            _df = _df[[c for c in _cols if c in _df.columns]]
                            st.session_state.results_df = _df
                            st.session_state.brand_name = _bi
                            st.session_state.scraping_done = True
                            _save_results_backup(_df, _bi)
                    results = []
                    for bi, bname in enumerate(_ss_suggest_brands):
                        site_status.text(f"🔴 [{bi+1}/{len(_ss_suggest_brands)}] ブランド「{bname}」を検索中...")
                        site_progress.progress(bi / len(_ss_suggest_brands))
                        burl = config["build_url"](bname)
                        purls = config["get_urls"](burl, max_pages, lambda msg, _s=site_status, _n=bname: _s.text(f"🔴 {_n}: {msg}"), save_callback=_ms_ss_save_cb)
                        if purls:
                            brand_results = [config["get_detail"](u) for u in purls]
                            brand_results = [r for r in brand_results if r]
                            results.extend(brand_results)
                    site_progress.progress(1.0)
                    for r in results:
                        r["サイト"] = site_name
                    all_results.extend(results)
                    # 中間保存（クラッシュ対策: 各サイト完了時にバックアップ）
                    if all_results:
                        _interim_df = pd.DataFrame(all_results)
                        _interim_cols = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                        _interim_df = _interim_df[[c for c in _interim_cols if c in _interim_df.columns]]
                        st.session_state.results_df = _interim_df
                        st.session_state.brand_name = brand_input
                        st.session_state.scraping_done = True
                        _save_results_backup(_interim_df, brand_input)
                    site_status.text(f"✅ {site_name}: {len(results)}件取得完了（{len(_ss_suggest_brands)}ブランド）")
                    continue
                base_url = config["build_url"](brand_input)
                site_status.text(f"🔗 {site_name}: {base_url}")
                def update_site_status(msg, _status=site_status, _icon=site_icon, _name=site_name):
                    _status.text(f"{_icon} {_name}: {msg}")
                # セカスト: 10ページごとに中間保存するコールバック
                _ss_save_cb = None
                if site_name == "セカスト":
                    def _ss_save_cb(cache_dict, _site=site_name, _bi=brand_input):
                        _rescued = list(cache_dict.values())
                        for _r in _rescued:
                            _r.setdefault("サイト", _site)
                        _all = all_results + _rescued
                        _df = pd.DataFrame(_all)
                        _cols = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                        _df = _df[[c for c in _cols if c in _df.columns]]
                        st.session_state.results_df = _df
                        st.session_state.brand_name = _bi
                        st.session_state.scraping_done = True
                        _save_results_backup(_df, _bi)
                    product_urls = config["get_urls"](base_url, max_pages, update_site_status, save_callback=_ss_save_cb)
                else:
                    product_urls = config["get_urls"](base_url, max_pages, update_site_status)
                if not product_urls:
                    site_status.text(f"⚠️ {site_name}: 商品が見つかりませんでした")
                    site_progress.progress(1.0)
                    continue
                site_status.text(f"📦 {site_name}: {len(product_urls)}件の商品を発見")
                if config.get("cached"):
                    site_status.text(f"{site_icon} {site_name}: データ整理中...")
                    results = [config["get_detail"](url) for url in product_urls]
                    results = [r for r in results if r]
                elif config["parallel"]:
                    def update_site_progress(completed, total, _bar=site_progress, _status=site_status, _name=site_name, _icon=site_icon):
                        _bar.progress(completed / total)
                        _status.text(f"{_icon} {_name}: [{completed}/{total}] 並列取得中...")
                    results = get_products_parallel(product_urls, config["get_detail"], max_workers=10, progress_callback=update_site_progress)
                else:
                    results = []
                    total = len(product_urls)
                    for i, url in enumerate(product_urls):
                        site_status.text(f"{site_icon} {site_name}: [{i+1}/{total}] 取得中...")
                        site_progress.progress((i + 1) / total)
                        data = config["get_detail"](url)
                        if data:
                            results.append(data)
                        time.sleep(random.uniform(0.3, 0.7))
                for r in results:
                    r["サイト"] = site_name
                all_results.extend(results)
                # 中間保存（クラッシュ対策: 各サイト完了時にバックアップ）
                if all_results:
                    _interim_df = pd.DataFrame(all_results)
                    _interim_cols = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                    _interim_df = _interim_df[[c for c in _interim_cols if c in _interim_df.columns]]
                    st.session_state.results_df = _interim_df
                    st.session_state.brand_name = brand_input
                    st.session_state.scraping_done = True
                    _save_results_backup(_interim_df, brand_input)
                site_status.text(f"✅ {site_name}: {len(results)}件取得完了")
                site_progress.progress(1.0)
            except Exception as e:
                st.error(f"⚠️ {site_name} のスクレイピング中にエラーが発生しました: {e}")
                # セカストはキャッシュから部分データを救出
                if site_name == "セカスト" and _secondstreet_cache:
                    rescued = list(_secondstreet_cache.values())
                    for r in rescued:
                        r["サイト"] = site_name
                    all_results.extend(rescued)
                    st.warning(f"⚠️ {site_name}: エラー前に取得済みの {len(rescued)}件 を保存しました")
                    # 中間保存
                    if all_results:
                        _interim_df = pd.DataFrame(all_results)
                        _interim_cols = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                        _interim_df = _interim_df[[c for c in _interim_cols if c in _interim_df.columns]]
                        st.session_state.results_df = _interim_df
                        st.session_state.brand_name = brand_input
                        st.session_state.scraping_done = True
                        _save_results_backup(_interim_df, brand_input)
                continue
        overall_progress.progress(1.0)
        overall_status.text(f"✅ {total_sites}サイトのスクレイピング完了！")
        _need_rerun = False
        if all_results:
            df = pd.DataFrame(all_results)
            columns = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
            df = df[[c for c in columns if c in df.columns]]
            st.session_state.results_df = df
            st.session_state.brand_name = brand_input
            st.session_state.scraping_done = True
            _save_results_backup(df, brand_input)
            _need_rerun = True
        else:
            st.warning("⚠️ どのサイトからも商品を取得できませんでした")
        if _need_rerun:
            st.rerun()
# 結果表示（session_stateから）
if st.session_state.scraping_done and st.session_state.results_df is not None:
    df = st.session_state.results_df
    st.success(f"✅ {len(df)}件の商品データを取得しました")
    if "サイト" in df.columns:
        site_counts = df["サイト"].value_counts()
        site_summary = " / ".join([f"{EC_SITES[s]['icon']} {s}: {c}件" for s, c in site_counts.items() if s in EC_SITES])
        st.info(f"📊 サイト別内訳: {site_summary}")
    # セカスト再スクレイピングボタン（不足分追加取得）
    _has_ss = ("サイト" in df.columns and "セカスト" in df["サイト"].values) or (st.session_state.get("selected_ec") == "セカスト")
    if _has_ss and st.session_state.brand_name:
        st.divider()
        _ss_count = df[df["サイト"] == "セカスト"].shape[0] if "サイト" in df.columns else len(df)
        st.warning(f"⚠️ セカスト現在 **{_ss_count}件** — 不足がある場合は再取得してください")
        if st.button("🔄 セカスト不足分を再スクレイピング", use_container_width=True, key="resume_ss_btn",
                      type="primary", help="取りこぼしたデータを追加取得します（既存データは保持）"):
            _resume_brand = st.session_state.brand_name
            # 既存セカストURLを取得
            if "サイト" in df.columns:
                _existing_ss_urls = set(df[df["サイト"] == "セカスト"]["URL"].dropna().tolist())
            else:
                _existing_ss_urls = set(df["URL"].dropna().tolist())
            _resume_status = st.empty()
            _resume_progress = st.progress(0)
            _resume_status.text(f"🔄 セカスト再スクレイピング中...（既存{len(_existing_ss_urls)}件をスキップ）")
            # サジェスト候補を使う
            _ss_brands = [_resume_brand]
            if (st.session_state.brand_suggest_results
                    and st.session_state.brand_suggest_keyword.lower() in _resume_brand.lower()):
                _ss_brands = [b["name"] for b in st.session_state.brand_suggest_results]
            _new_results = []
            for _bi, _bname in enumerate(_ss_brands):
                _resume_status.text(f"🔄 [{_bi+1}/{len(_ss_brands)}] {_bname} を再スクレイピング中...")
                _burl = secondstreet_build_url(_bname, "")
                def _resume_save_cb(cache_dict):
                    pass  # 中間保存は不要（マージ時に一括保存）
                _purls = secondstreet_get_product_urls(
                    _burl, 999,
                    progress_callback=lambda msg, _s=_resume_status, _n=_bname: _s.text(f"🔄 {_n}: {msg}"),
                    save_callback=_resume_save_cb,
                    existing_urls=_existing_ss_urls,
                )
                if _purls:
                    _brand_results = [secondstreet_get_product_detail(u) for u in _purls]
                    _brand_results = [r for r in _brand_results if r]
                    _new_results.extend(_brand_results)
                _resume_progress.progress((_bi + 1) / len(_ss_brands))
            if _new_results:
                for r in _new_results:
                    r.setdefault("サイト", "セカスト")
                _new_df = pd.DataFrame(_new_results)
                # 既存DFとマージ（URL重複除去）
                _merged = pd.concat([df, _new_df], ignore_index=True)
                _merged = _merged.drop_duplicates(subset=["URL"], keep="first")
                _cols = ["サイト", "ブランド", "商品名", "通称", "カテゴリ", "品番", "型番", "価格", "参考上代", "ランク", "サイズ", "実寸サイズ", "カラー", "素材", "性別", "製造国", "付属品", "URL"]
                _merged = _merged[[c for c in _cols if c in _merged.columns]]
                st.session_state.results_df = _merged
                _save_results_backup(_merged, _resume_brand)
                _resume_status.text(f"✅ {len(_new_results)}件を追加取得（合計{len(_merged)}件）")
                st.rerun()
            else:
                _resume_status.text("ℹ️ 新規データなし（全件取得済みの可能性）")
                _resume_progress.progress(1.0)
    # メトリクス（2行x2列でスマホでも見やすく）
    mcol1, mcol2 = st.columns(2)
    with mcol1:
        st.metric("取得件数", f"{len(df)}件")
    with mcol2:
        if "価格" in df.columns and df["価格"].notna().any():
            st.metric("平均価格", f"¥{df['価格'].mean():,.0f}")
    mcol3, mcol4 = st.columns(2)
    with mcol3:
        if "価格" in df.columns and df["価格"].notna().any():
            st.metric("最安値", f"¥{df['価格'].min():,.0f}")
    with mcol4:
        if "価格" in df.columns and df["価格"].notna().any():
            st.metric("最高値", f"¥{df['価格'].max():,.0f}")
    st.dataframe(df, use_container_width=True, height=400)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if "サイト" in df.columns:
        filename = f"allsites_{st.session_state.brand_name}_{timestamp}.xlsx"
    else:
        ec_name = st.session_state.get("selected_ec", "result").lower().replace(" ", "")
        filename = f"{ec_name}_{st.session_state.brand_name}_{timestamp}.xlsx"
    excel_data = to_excel(df)
    st.download_button(
        label="📥 Excelダウンロード",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.xlsx",
        use_container_width=True
    )
    # ===== 品番→正式名称検索セクション =====
    has_hinban = "品番" in df.columns or "型番" in df.columns
    if has_hinban:
        st.divider()
        st.markdown("### 🔎 品番→正式名称検索")
        unique_hinban = extract_unique_hinban(df)
        if not unique_hinban:
            st.info("📝 有効な品番が見つかりませんでした")
        else:
            st.info(f"📝 ユニーク品番数: **{len(unique_hinban)}**件（名寄せ済み）")
            with st.expander("品番一覧を確認"):
                st.write(", ".join(unique_hinban[:100]))
                if len(unique_hinban) > 100:
                    st.caption(f"...他 {len(unique_hinban) - 100}件")
            wear_search_enabled = st.checkbox(
                "👗 WEAR検索で補完（ZOZOTOWN系ブランドに有効）",
                value=True,
                key="hinban_wear_search",
                help="WEAR APIで品番を検索。TODAYFUL等のZOZOTOWN取扱ブランドに高精度。APIキー不要・高速。",
            )
            web_search_enabled = st.checkbox(
                "🌐 Web検索で補完（DDG/Brave、レート制限あり）",
                value=False,
                key="hinban_web_search",
                help="WEARでも見つからなかった品番をDDG/Braveで検索。レート制限があるため数件のみ。",
            )
            max_lookup = st.slider(
                "対象品番数の上限",
                min_value=1,
                max_value=len(unique_hinban),
                value=len(unique_hinban),
                key="hinban_max_lookup",
            )
            lookup_button = st.button("🔍 正式名称を取得", type="primary", use_container_width=True, key="hinban_lookup_btn")
            if lookup_button:
                progress_bar = st.progress(0)
                status_text = st.empty()
                def update_lookup_progress(current, total, msg):
                    if total > 0:
                        progress_bar.progress(min(current / total, 1.0))
                    status_text.text(msg)
                mapping_df = lookup_official_names(
                    df,
                    brand=st.session_state.brand_name,
                    max_lookup=max_lookup,
                    wear_search=wear_search_enabled,
                    web_search=web_search_enabled,
                    progress_callback=update_lookup_progress,
                )
                st.session_state.hinban_mapping_df = mapping_df
                st.session_state.hinban_lookup_done = True
                st.rerun()
        if st.session_state.hinban_lookup_done and st.session_state.hinban_mapping_df is not None:
            mapping_df = st.session_state.hinban_mapping_df
            found_count = (mapping_df["正式名称"] != "").sum()
            local_count = (mapping_df["ソース"] == "ローカル").sum()
            wear_count = (mapping_df["ソース"] == "WEAR").sum()
            web_count = (mapping_df["ソース"] == "Web検索").sum()
            source_detail = f"（ローカル: {local_count}件"
            if wear_count > 0:
                source_detail += f" / WEAR: {wear_count}件"
            if web_count > 0:
                source_detail += f" / Web: {web_count}件"
            source_detail += "）"
            st.success(f"✅ {found_count}/{len(mapping_df)}件の品番で正式名称を特定 {source_detail}")
            enriched_df, filtered_df = enrich_mapping_with_kde(df, mapping_df, min_price=10000)
            excluded_count = len(enriched_df[
                (enriched_df["正式名称"] != "") &
                (enriched_df["KDE代表価格"] > 0) &
                (enriched_df["KDE代表価格"] <= 10000)
            ])
            no_name_count = (enriched_df["正式名称"] == "").sum()
            if excluded_count > 0:
                st.info(f"📊 KDE分析: **{excluded_count}件**をKDE代表価格≤¥10,000で除外 / 名称未特定: {no_name_count}件")
            display_cols = ["品番", "正式名称", "件数", "KDE代表価格", "価格レンジ", "ソース"]
            display_df = filtered_df[[c for c in display_cols if c in filtered_df.columns]].copy()
            if "KDE代表価格" in display_df.columns:
                display_df["KDE代表価格"] = display_df["KDE代表価格"].apply(
                    lambda x: f"¥{x:,.0f}" if x > 0 else ""
                )
            st.dataframe(display_df, use_container_width=True, height=300)
            if excluded_count > 0:
                with st.expander(f"🚫 除外された品番を確認（{excluded_count}件: KDE代表価格≤¥10,000）"):
                    excluded_df = enriched_df[
                        (enriched_df["正式名称"] != "") &
                        (enriched_df["KDE代表価格"] > 0) &
                        (enriched_df["KDE代表価格"] <= 10000)
                    ][[c for c in display_cols if c in enriched_df.columns]].copy()
                    if "KDE代表価格" in excluded_df.columns:
                        excluded_df["KDE代表価格"] = excluded_df["KDE代表価格"].apply(
                            lambda x: f"¥{x:,.0f}" if x > 0 else ""
                        )
                    st.dataframe(excluded_df, use_container_width=True)
            st.divider()
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            brand = st.session_state.brand_name
            def build_single_sheet_excel():
                """1シートに品番マッピング＋KDEと正式名称集計を縦に並べて出力"""
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = Workbook()
                ws = wb.active
                ws.title = "分析結果"
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, size=11, color="FFFFFF")
                section_font = Font(bold=True, size=13)
                section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'),
                )
                row = 1
                # ===== 上段: 品番マッピング＋KDE =====
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                c = ws.cell(row=row, column=1, value="■ 品番マッピング＋KDE")
                c.font = section_font; c.fill = section_fill
                row += 1
                s1_headers = ["ブランド名", "品番", "KDE代表価格", "件数", "ソース", "価格レンジ"]
                for ci, h in enumerate(s1_headers, 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font = header_font; c.fill = header_fill; c.border = thin_border
                    c.alignment = Alignment(horizontal='center')
                row += 1
                for _, dr in filtered_df.iterrows():
                    s1_row_data = {
                        "ブランド名": brand.upper(),
                        "品番": dr.get("品番", ""),
                        "KDE代表価格": dr.get("KDE代表価格", ""),
                        "件数": dr.get("件数", ""),
                        "ソース": dr.get("ソース", ""),
                        "価格レンジ": dr.get("価格レンジ", ""),
                    }
                    for ci, col in enumerate(s1_headers, 1):
                        val = s1_row_data[col]
                        if col == "KDE代表価格" and val and float(val) > 0:
                            val = int(float(val))
                        c = ws.cell(row=row, column=ci, value=val)
                        c.border = thin_border
                        if col in ("件数", "KDE代表価格"):
                            c.alignment = Alignment(horizontal='right')
                            if col == "KDE代表価格":
                                c.number_format = '#,##0'
                    row += 1
                row += 2
                # ===== 下段: 正式名称集計 =====
                original_df = st.session_state.results_df.copy()
                name_map = dict(zip(filtered_df["品番"], filtered_df["正式名称"]))
                merged = original_df.copy()
                merged["正式名称"] = ""
                if "品番" in merged.columns:
                    merged["正式名称"] = merged["品番"].map(
                        lambda x: name_map.get(str(x).strip(), "") if pd.notna(x) else ""
                    )
                if "型番" in merged.columns:
                    for idx, mrow in merged.iterrows():
                        if not mrow.get("正式名称"):
                            kataban = str(mrow.get("型番", "")).strip()
                            for part in kataban.split(" / "):
                                part = part.strip()
                                if part in name_map and name_map[part]:
                                    merged.at[idx, "正式名称"] = name_map[part]
                                    break
                named = merged[merged["正式名称"] != ""].copy()
                groups = []
                if len(named) > 0 and "価格" in named.columns:
                    for name, grp in named.groupby("正式名称"):
                        prices = pd.to_numeric(grp["価格"], errors="coerce").dropna().tolist()
                        prices = [p for p in prices if p > 0]
                        if not prices:
                            continue
                        kde_price = _kde_peak_price(prices)
                        groups.append({
                            "ブランド": brand.upper(),
                            "正式名称": name,
                            "件数": len(prices),
                            "KDE代表価格": int(kde_price),
                            "最安値": int(min(prices)),
                            "最高値": int(max(prices)),
                        })
                    groups.sort(key=lambda x: -x["KDE代表価格"])
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                c = ws.cell(row=row, column=1, value="■ 正式名称集計（同名統合・KDE価格順）")
                c.font = section_font; c.fill = section_fill
                row += 1
                s2_headers = ["ブランド名", "正式名称", "KDE代表価格", "件数"]
                for ci, h in enumerate(s2_headers, 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font = header_font; c.fill = header_fill; c.border = thin_border
                    c.alignment = Alignment(horizontal='center')
                row += 1
                for g in groups:
                    s2_row_data = {
                        "ブランド名": g.get("ブランド", ""),
                        "正式名称": g.get("正式名称", ""),
                        "KDE代表価格": g.get("KDE代表価格", ""),
                        "件数": g.get("件数", ""),
                    }
                    for ci, col in enumerate(s2_headers, 1):
                        val = s2_row_data[col]
                        c = ws.cell(row=row, column=ci, value=val)
                        c.border = thin_border
                        if col in ("KDE代表価格", "件数"):
                            c.alignment = Alignment(horizontal='right')
                            if col == "KDE代表価格":
                                c.number_format = '#,##0'
                    row += 1
                ws.column_dimensions['A'].width = 18
                ws.column_dimensions['B'].width = 48
                ws.column_dimensions['C'].width = 16
                ws.column_dimensions['D'].width = 10
                ws.column_dimensions['E'].width = 14
                ws.column_dimensions['F'].width = 20
                output = BytesIO()
                wb.save(output)
                return output.getvalue()
            excel_data = build_single_sheet_excel()
            st.download_button(
                label="📥 Excelダウンロード",
                data=excel_data,
                file_name=f"{brand}_analysis_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.xlsx",
                use_container_width=True,
                key="hinban_mapping_dl",
            )
# 使い方
with st.expander("📖 使い方"):
    st.markdown("""
    ### 使い方
    1. **ECサイトを選択**: コメ兵 / RAGTAG / トレファク / セカスト
    2. **ブランド名を入力**: 英語・スペースなし（例: `fendi`, `louisvuitton`）
    3. **カテゴリを選択**: 絞り込む場合は選択
    4. **取得ページ数を設定**: 1ページ約50〜90件
    5. **「スクレイピング開始」をクリック**
    6. 完了後「Excelダウンロード」で保存
    ### ブランド名の書き方
    | ブランド | コメ兵 | RAGTAG | トレファク | セカスト |
    |----------|--------|--------|------------|---------|
    | フェンディ | `fendi` | `FENDI` | `fendi` | `fendi` |
    | グッチ | `gucci` | `GUCCI` | `gucci` | `gucci` |
    | ルイヴィトン | `louisvuitton` | `LOUISVUITTON` | `louis vuitton` | `louis vuitton` |
    | シャネル | `chanel` | `CHANEL` | `chanel` | `chanel` |
    | エルメス | `hermes` | `HERMES` | `hermes` | `hermes` |
    | プラダ | `prada` | `PRADA` | `prada` | `prada` |
    | セリーヌ | `celine` | `CELINE` | `celine` | `celine` |
    | ロエベ | `loewe` | `LOEWE` | `loewe` | `loewe` |
    ※RAGTAGは自動で大文字変換されます
    ⚠️ **セカスト利用時の注意**:
    - 他のサイトと同様にrequestsベースで動作します
    """)
st.divider()
st.caption("⚠️ 利用は自己責任で。robots.txt・利用規約を確認してください。")
